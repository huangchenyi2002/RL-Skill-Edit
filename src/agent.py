# src/agent.py
"""
Student Agent
修改：
  1. _detect_activation 增强，兼容 Claude 系列
     四层检测：精确匹配→Claude表达→关键词→代码推断
  2. step_execution_signals 记录（β₂/β₃ signal density）
  3. run_task_no_skill()（π₀ 基线）
  4. Tier 2 注入时明确引导 Student 照着代码示例写
  5. exec_cache 线程安全
"""

import hashlib
import os
import platform
import re
import shutil
import subprocess
import tempfile
import threading
from dataclasses import dataclass, field
from src.client import OpenRouterClient
from src.skill_library import SkillLibrary


@dataclass
class TrajectoryStep:
    step:            int
    observation:     str
    action:          str
    external_result: str
    activated_skill: str
    activation_mode: str
    system_prompt:   str   = ""
    tokens_used:     int   = 0
    cost_usd:        float = 0.0
    # Parser LLM routing onto the frozen Skill-history universe.
    # None means mapping was not performed; [] is an explicit UNASSIGNED.
    applicable_history_ids: list[str] | None = None
    history_mapping_status: str = "missing"
    history_mapping_reason: str = ""


@dataclass
class Trajectory:
    task_id:          str
    task_description: str
    steps:            list[TrajectoryStep] = field(
        default_factory=list)
    final_reward:     float = 0.0
    success:          bool  = False
    total_tokens:     int   = 0
    total_cost_usd:   float = 0.0
    activated_skills: list[str] = field(
        default_factory=list)
    score_detail:     dict = field(
        default_factory=dict)
    step_execution_signals: list[dict] = field(
        default_factory=list)
    evaluation_valid: bool = True
    invalid_reason: str = ""
    hard_reward: float = 0.0
    soft_reward: float = 0.0
    attempt_scores: list[float] = field(default_factory=list)


class StudentAgent:

    def __init__(
        self,
        config: dict,
        client: OpenRouterClient,
    ):
        self.config      = config
        self.client      = client
        self.model       = config["student"]["model"]
        self.temp        = config["student"]["temperature"]
        self.max_tok     = config["student"]["max_tokens"]
        self.max_steps   = config["student"]["max_steps"]
        self._exec_cache: dict = {}
        self._cache_lock = threading.Lock()
        # subprocess 并发限制：4（稳定优先）
        # Excel 执行本身够快，多并发也没意义
        self._exec_semaphore = threading.Semaphore(4)

    # ──────────────────────────────────────────────
    # 主入口
    # ──────────────────────────────────────────────

    def run_task(
        self,
        task:            dict,
        library:         SkillLibrary,
        activation_mode: str = "implicit",
        forced_skill_id: str = None,
        verifier_feedback: bool = True,
        expose_answer_metadata: bool = True,
        seed: int = None,
    ) -> Trajectory:
        traj = Trajectory(
            task_id          = task["task_id"],
            task_description = task["description"],
        )

        current_skill_id = (
            forced_skill_id
            if activation_mode == "harness"
            else None
        )

        system = self._build_system(
            library         = library,
            activation_mode = activation_mode,
            forced_skill_id = forced_skill_id,
            task            = task,
            expose_answer_metadata=expose_answer_metadata,
        )
        messages = [
            {"role": "user",
             "content": task["description"]}
        ]

        spreadsheet  = task.get("spreadsheet", {})
        init_file    = spreadsheet.get("init_file", "")
        golden_file  = spreadsheet.get("golden_file", "")
        answer_pos   = spreadsheet.get("answer_position", "")
        answer_sheet = spreadsheet.get("answer_sheet", "")
        has_golden   = (
            golden_file and os.path.exists(golden_file)
            and init_file and os.path.exists(init_file)
            and answer_pos
        )

        for step_idx in range(self.max_steps):

            # Tier 2 注入 execution_body
            full_system = system
            if current_skill_id and library:
                sk = library.get(current_skill_id)
                if sk and sk.execution_body:
                    full_system = (
                        system
                        + "\n\n" + "=" * 50 + "\n"
                        + f"[ACTIVE SKILL: {sk.name}]\n"
                        + "You MUST follow this skill.\n"
                        + "=" * 50 + "\n\n"
                        + sk.execution_body
                        + "\n\n" + "=" * 50 + "\n"
                        + "MANDATORY — follow the skill "
                        + "above, WITH THESE OVERRIDES:\n"
                        + "- Use the Code Examples as "
                        + "your primary templates\n"
                        + "- Follow EVERY rule listed\n"
                        + "- Adapt the patterns to this "
                        + "specific task\n"
                        + "⚠️ **VARIABLE NAMES**: "
                        + "the execution environment "
                        + "provides ONLY `wb_path`. "
                        + "Do NOT use `INPUT_PATH`, "
                        + "`OUTPUT_PATH`, `IN_PATH` etc. "
                        + "Always: "
                        + "`wb = load_workbook(wb_path)` "
                        + "and `wb.save(wb_path)`\n"
                        + "- ALWAYS end with "
                        + "wb.save(wb_path)\n"
                        + "=" * 50
                    )

            response, usage = self.client.chat(
                model       = self.model,
                messages    = messages,
                system      = full_system,
                temperature = self.temp,
                max_tokens  = self.max_tok,
                call_type   = "student_rollout",
                seed=(None if seed is None else int(seed) + step_idx),
            )
            traj.total_tokens   += usage["total_tokens"]
            traj.total_cost_usd += usage["cost_usd"]
            if not usage.get("ok", True):
                traj.evaluation_valid = False
                traj.invalid_reason = (
                    f"api:{usage.get('error_kind', 'unknown')}")

            # 检测 Skill 激活
            activated, sid = self._detect_activation(
                response, library,
                activation_mode, forced_skill_id,
            )
            if activated and sid and not current_skill_id:
                current_skill_id = sid
            # harness 模式：current_skill_id 已被预设
            # 确保 activated_skills 被记录
            if current_skill_id and \
                    current_skill_id not in \
                    traj.activated_skills:
                traj.activated_skills.append(
                    current_skill_id)

            traj.steps.append(TrajectoryStep(
                step            = step_idx,
                observation     = messages[-1]["content"],
                action          = response,
                external_result = "",
                activated_skill = current_skill_id or "0",
                activation_mode = activation_mode,
                system_prompt   = full_system,
                tokens_used     = usage["total_tokens"],
                cost_usd        = usage["cost_usd"],
            ))

            # 多步对话 + step signal 记录
            if has_golden:
                code = self._extract_code(response)
                if code:
                    trial_score, trial_detail = \
                        self._execute_and_score(
                            code         = code,
                            init_file    = init_file,
                            golden_file  = golden_file,
                            answer_pos   = answer_pos,
                            answer_sheet = answer_sheet,
                        )

                    err = trial_detail.get("error", "")
                    traj.step_execution_signals.append({
                        "step":         step_idx,
                        "code_snippet": code[:300],
                        "error":        err,
                        "matched":      trial_detail.get(
                            "matched", 0),
                        "total":        trial_detail.get(
                            "total", 0),
                        "score":        trial_score,
                        "has_error":    bool(err),
                        "has_verified": False,
                    })

                    if trial_score >= 0.8:
                        break

                    if (verifier_feedback
                            and step_idx < self.max_steps - 1):
                        messages.append(
                            {"role": "assistant",
                             "content": response})
                        if err:
                            feedback = (
                                f"Your code produced "
                                f"an error:\n{err}\n\n"
                                "Please fix the code "
                                "following the skill's "
                                "Code Examples as "
                                "templates.\n")
                        else:
                            sc = f"{trial_score:.1f}"
                            feedback = (
                                "Your code ran but "
                                "produced wrong results"
                                f" (score={sc}).\n"
                                "The output cells did "
                                "not match expected "
                                "values.\n"
                                "Please review your "
                                "logic carefully and "
                                "fix the code.\n")
                        messages.append(
                            {"role": "user",
                             "content": (
                                 feedback +
                                 "Make sure to:\n"
                                 "1. Load: wb = "
                                 "load_workbook(wb_path)\n"
                                 "2. Apply all changes\n"
                                 "3. Save: wb.save(wb_path)\n\n"
                                 "Provide complete fixed "
                                 "code in a ```python block."
                             )})
                        continue
                break
            else:
                if activation_mode == "implicit":
                    break
                else:
                    if step_idx >= self.max_steps - 1:
                        break
                    messages.append(
                        {"role": "assistant",
                         "content": response})

        final_reward, score_detail = \
            self._compute_reward(
                task             = task,
                steps            = traj.steps,
                activation_mode  = activation_mode,
                activated_skills = traj.activated_skills,
                current_skill_id = current_skill_id,
            )
        traj.final_reward = final_reward
        traj.soft_reward  = final_reward
        traj.hard_reward  = 1.0 if final_reward >= 0.8 else 0.0
        traj.success      = final_reward > 0.5
        traj.score_detail = score_detail
        if score_detail.get("valid_for_estimation") is False:
            traj.evaluation_valid = False
            traj.invalid_reason = score_detail.get(
                "reason", "invalid_reward_artifact")
        traj.score_detail["answer_pos"]   = answer_pos
        traj.score_detail["answer_sheet"] = answer_sheet
        return traj

    # ──────────────────────────────────────────────
    # π₀ 基线：无 skill 模式
    # ──────────────────────────────────────────────

    def run_task_no_skill(self, task: dict) -> Trajectory:
        traj = Trajectory(
            task_id          = task["task_id"],
            task_description = task["description"],
        )

        spreadsheet  = task.get("spreadsheet", {})
        init_file    = spreadsheet.get("init_file", "")
        golden_file  = spreadsheet.get("golden_file", "")
        answer_pos   = spreadsheet.get("answer_position", "")
        answer_sheet = spreadsheet.get("answer_sheet", "")

        info = ""
        if answer_pos:
            info = (
                f"\nTarget sheet: {answer_sheet}"
                f"\nTarget range: {answer_pos}")

        system = (
            "You are an AI agent solving Excel tasks.\n"
            "Write Python code to solve the task.\n"
            "CRITICAL REQUIREMENTS:\n"
            "1. Workbook path: wb_path (already bound)\n"
            "2. Your code MUST:\n"
            "   a) wb = load_workbook(wb_path)\n"
            "   b) Make required modifications\n"
            "   c) wb.save(wb_path)  <- REQUIRED\n"
            "3. Do NOT just print results\n"
            f"{info}\n\n"
            "Code template:\n"
            "```python\n"
            "from openpyxl import load_workbook\n"
            "import pandas as pd\n\n"
            "wb = load_workbook(wb_path)\n"
            "ws = wb.active\n"
            "# Your solution here\n"
            "wb.save(wb_path)  # REQUIRED\n"
            "```"
        )

        messages = [
            {"role": "user",
             "content": task["description"]}
        ]

        has_golden = (
            golden_file
            and os.path.exists(golden_file)
            and init_file
            and os.path.exists(init_file)
            and answer_pos)

        # multi-step retry（和 run_task 一致）
        max_steps = self.max_steps
        final_score = 0.0
        final_detail = {"error": "no_run"}

        for step_idx in range(max_steps):
            response, usage = self.client.chat(
                model       = self.model,
                messages    = messages,
                system      = system,
                temperature = self.temp,
                max_tokens  = self.max_tok,
                call_type   = "student_rollout",
            )
            traj.total_tokens   += usage[
                "total_tokens"]
            traj.total_cost_usd += usage["cost_usd"]
            if not usage.get("ok", True):
                traj.evaluation_valid = False
                traj.invalid_reason = (
                    f"api:{usage.get('error_kind', 'unknown')}")

            traj.steps.append(TrajectoryStep(
                step            = step_idx,
                observation     = messages[-1][
                    "content"],
                action          = response,
                external_result = "",
                activated_skill = "0",
                activation_mode = "no_skill",
                system_prompt   = system,
                tokens_used     = usage[
                    "total_tokens"],
                cost_usd        = usage["cost_usd"],
            ))

            if not has_golden:
                final_detail = {"error": "no_golden"}
                break

            code = self._extract_code(response)
            if not code:
                final_detail = {"error": "no_code_extracted"}
                break

            score, detail = self._execute_and_score(
                code         = code,
                init_file    = init_file,
                golden_file  = golden_file,
                answer_pos   = answer_pos,
                answer_sheet = answer_sheet,
            )
            traj.attempt_scores.append(score)
            if score >= final_score:
                final_score = score
                final_detail = detail

            if score >= 0.8:
                break

            # 失败 → 反馈重试
            if step_idx < max_steps - 1:
                err = detail.get("error", "")
                fb = (f"Error:\n{err}\n\nFix the code."
                      if err else
                      f"Code ran but wrong results "
                      f"(score={score:.2f}). "
                      "Review logic and fix.")
                messages.append(
                    {"role": "assistant",
                     "content": response})
                messages.append(
                    {"role": "user",
                     "content": (
                         fb + "\n"
                         "wb = load_workbook(wb_path)\n"
                         "wb.save(wb_path)")})

        traj.final_reward = final_score
        traj.soft_reward = final_score
        traj.hard_reward = 1.0 if final_score >= 0.8 else 0.0
        traj.score_detail = final_detail
        if final_detail.get("valid_for_estimation") is False:
            traj.evaluation_valid = False
            traj.invalid_reason = final_detail.get(
                "reason", "invalid_reward_artifact")
        traj.success = final_score > 0.5
        return traj

    # ──────────────────────────────────────────────
    # Teacher rollout：用指定模型执行任务
    # ──────────────────────────────────────────────

    def run_task_with_model(
        self,
        task: dict,
        library: SkillLibrary,
        model_override: str = None,
        activation_mode: str = "harness",
        forced_skill_id: str = None,
        max_steps_override: int = None,
    ) -> Trajectory:
        """
        用指定模型（如 Teacher）执行任务。
        包含多步重试（出错 → feedback → 重试）。
        用于 §3.2 步骤 5 的 Teacher rollout。
        """
        model = model_override or self.model
        traj = Trajectory(
            task_id=task["task_id"],
            task_description=task["description"],
        )

        spreadsheet = task.get("spreadsheet", {})
        init_file = spreadsheet.get("init_file", "")
        golden_file = spreadsheet.get(
            "golden_file", "")
        answer_pos = spreadsheet.get(
            "answer_position", "")
        answer_sheet = spreadsheet.get(
            "answer_sheet", "")

        system = self._build_system(
            library=library,
            activation_mode=activation_mode,
            forced_skill_id=forced_skill_id,
            task=task,
        )

        # Tier 2 注入
        if forced_skill_id and library:
            sk = library.get(forced_skill_id)
            if sk and sk.execution_body:
                system = (
                    system
                    + "\n\n" + "=" * 50 + "\n"
                    + f"[ACTIVE SKILL: {sk.name}]\n"
                    + "You MUST follow this skill.\n"
                    + "=" * 50 + "\n\n"
                    + sk.execution_body
                    + "\n\n" + "=" * 50 + "\n"
                    + "MANDATORY — follow the skill "
                    + "above.\n"
                    + "=" * 50
                )

        if forced_skill_id:
            traj.activated_skills.append(
                forced_skill_id)

        has_golden = (
            golden_file
            and os.path.exists(golden_file)
            and init_file
            and os.path.exists(init_file)
            and answer_pos)

        messages = [
            {"role": "user",
             "content": task["description"]}
        ]

        # 多步循环（和 run_task 一致）
        max_steps = (max_steps_override
                     if max_steps_override
                     else self.max_steps)
        # Teacher 用小 max_tokens 加速
        # (2048 token = ~1500 行代码，绰绰有余)
        teacher_max_tok = min(self.max_tok, 2048)
        for step_idx in range(max_steps):
            response, usage = self.client.chat(
                model=model,
                messages=messages,
                system=system,
                temperature=0.0,
                max_tokens=teacher_max_tok,
                call_type="teacher_rollout",
            )
            traj.total_tokens += usage[
                "total_tokens"]
            traj.total_cost_usd += usage["cost_usd"]
            if not usage.get("ok", True):
                traj.evaluation_valid = False
                traj.invalid_reason = (
                    f"api:{usage.get('error_kind', 'unknown')}")

            traj.steps.append(TrajectoryStep(
                step=step_idx,
                observation=messages[-1]["content"],
                action=response,
                external_result="",
                activated_skill=(
                    forced_skill_id or "0"),
                activation_mode=activation_mode,
                system_prompt=system,
                tokens_used=usage["total_tokens"],
                cost_usd=usage["cost_usd"],
            ))

            if has_golden:
                code = self._extract_code(response)
                if code:
                    score, detail = \
                        self._execute_and_score(
                            code=code,
                            init_file=init_file,
                            golden_file=golden_file,
                            answer_pos=answer_pos,
                            answer_sheet=answer_sheet,
                        )

                    if score >= 0.8:
                        traj.final_reward = score
                        traj.soft_reward = score
                        traj.hard_reward = 1.0
                        traj.attempt_scores.append(score)
                        traj.score_detail = detail
                        traj.success = True
                        return traj

                    err = detail.get("error", "")
                    if step_idx < max_steps - 1:
                        messages.append(
                            {"role": "assistant",
                             "content": response})
                        if err:
                            fb = (f"Error:\n{err}\n\n"
                                  "Fix the code.")
                        else:
                            fb = (
                                f"Code ran but wrong "
                                f"results (score="
                                f"{score:.1f}). "
                                "Review logic and fix.")
                        messages.append(
                            {"role": "user",
                             "content": (
                                 fb + "\n"
                                 "wb = load_workbook("
                                 "wb_path)\n"
                                 "wb.save(wb_path)")})
                        continue
                break
            else:
                break

        # 最终评分
        final_reward, score_detail = \
            self._compute_reward(
                task=task,
                steps=traj.steps,
                activation_mode=activation_mode,
                activated_skills=(
                    traj.activated_skills),
                current_skill_id=forced_skill_id,
            )
        traj.final_reward = final_reward
        traj.soft_reward = final_reward
        traj.hard_reward = 1.0 if final_reward >= 0.8 else 0.0
        traj.success = final_reward > 0.5
        traj.score_detail = score_detail
        if score_detail.get("valid_for_estimation") is False:
            traj.evaluation_valid = False
            traj.invalid_reason = score_detail.get(
                "reason", "invalid_reward_artifact")
        return traj

    # ──────────────────────────────────────────────
    # 系统提示
    # ──────────────────────────────────────────────

    def _build_system(
        self,
        library:         SkillLibrary,
        activation_mode: str,
        forced_skill_id: str  = None,
        task:            dict = None,
        expose_answer_metadata: bool = True,
    ) -> str:
        info = ""
        if task and expose_answer_metadata:
            sp        = task.get("spreadsheet", {})
            ans_pos   = sp.get("answer_position", "")
            ans_sheet = sp.get("answer_sheet", "")
            if ans_pos:
                info = (
                    f"\nTarget sheet: {ans_sheet}"
                    f"\nTarget range: {ans_pos}")

        code_req = (
            "\n\nCRITICAL REQUIREMENTS:\n"
            "1. The Excel workbook is at: wb_path\n"
            "2. Your code MUST:\n"
            "   a) Load:  wb = load_workbook(wb_path)\n"
            "   b) Make all required modifications\n"
            "   c) Save:  wb.save(wb_path)  <- REQUIRED!\n"
            "3. Do NOT just print results\n"
            f"{info}\n"
            "\nCode template:\n"
            "```python\n"
            "from openpyxl import load_workbook\n"
            "import pandas as pd\n\n"
            "wb = load_workbook(wb_path)\n"
            "ws = wb.active\n\n"
            "# Your solution here\n\n"
            "wb.save(wb_path)  # REQUIRED!\n"
            "```"
        )

        if activation_mode == "harness" and forced_skill_id:
            sk   = library.get(forced_skill_id) \
                   if library else None
            name = sk.name if sk else forced_skill_id
            return (
                f"You are an AI agent solving Excel "
                f"tasks. You MUST use the skill "
                f"'{name}' provided below.\n\n"
                f"CRITICAL RULES:\n"
                f"- You MUST follow ALL rules and "
                f"patterns in the skill document.\n"
                f"- Use the skill's code examples "
                f"as your primary templates.\n"
                f"- Do NOT ignore the skill. Every "
                f"rule in the skill exists for a "
                f"reason.\n"
                f"- Adapt the skill's approach to "
                f"this specific task.\n"
                + code_req
            )
        elif activation_mode == "implicit":
            catalog = library.startup_catalog() \
                      if library else "(no skills)"
            return (
                "You are an AI agent solving "
                "Excel tasks.\n\n"
                "Available skills:\n"
                + catalog
                + "\n\nInstructions:\n"
                "1. Select the best skill by writing:\n"
                "   Using skill: <skill name>\n"
                "2. When activated, follow the skill's "
                "Code Examples as templates.\n"
                "3. Provide a complete Python solution.\n"
                + code_req
            )
        return (
            "You are an AI agent solving Excel tasks.\n"
            + code_req
        )

    # ──────────────────────────────────────────────
    # Skill 激活检测（增强版，兼容 Claude 系列）
    # 四层检测策略
    # ──────────────────────────────────────────────

    def _detect_activation(
        self,
        response:        str,
        library:         SkillLibrary,
        activation_mode: str,
        forced_skill_id: str,
    ) -> tuple[bool, str]:
        if activation_mode == "harness":
            return True, forced_skill_id
        if not library:
            return False, None

        resp_lower = response.lower()

        # ── 层1：精确匹配 "using skill: xxx" ─────
        pattern = (
            r"using skill[:\s]+"
            r"([a-zA-Z0-9_\-\s]+)"
        )
        match = re.search(pattern, resp_lower)
        if match:
            mentioned = match.group(1).strip().lower()
            # 精确匹配
            for sk in library:
                all_names = [sk.name.lower()] + [
                    n.lower() for n in
                    getattr(sk, "name_history", [])]
                if mentioned in all_names:
                    return True, sk.skill_id
            # 包含匹配
            for sk in library:
                all_names = [sk.name.lower()] + [
                    n.lower() for n in
                    getattr(sk, "name_history", [])]
                for nm in all_names:
                    if (mentioned in nm
                            or nm in mentioned):
                        return True, sk.skill_id

        # ── 层2：Claude 系列常见激活表达 ──────────
        # Claude 不说 "Using skill:"
        # 而说 "I'll use the X skill" 等
        claude_patterns = [
            r"i['\u2019ll ]+use (?:the )?([a-zA-Z0-9_\-\s]+?)(?:\s+skill)?[,\.]",
            r"applying (?:the )?([a-zA-Z0-9_\-\s]+?)(?:\s+skill)?[,\.]",
            r"(?:this|the) ([a-zA-Z0-9_\-\s]+?) skill",
            r"skill[:\s]+['\"]?([a-zA-Z0-9_\-\s]+?)['\"]?[\s,\.]",
            r"(?:will|should) use (?:the )?([a-zA-Z0-9_\-\s]+?)(?:\s+skill)?[,\.]",
            r"(?:selected|choosing|activate)[:\s]+([a-zA-Z0-9_\-\s]+?)(?:\s+skill)?[,\.\n]",
        ]
        for pat in claude_patterns:
            m = re.search(pat, resp_lower)
            if m:
                mentioned = m.group(1).strip().lower()
                if len(mentioned) < 3:
                    continue
                for sk in library:
                    all_names = [sk.name.lower()] + [
                        n.lower() for n in
                        getattr(sk, "name_history", [])]
                    for nm in all_names:
                        if (mentioned in nm
                                or nm in mentioned):
                            return True, sk.skill_id

        # ── 层3：skill 名称 / 历史名称关键词匹配 ──
        for sk in library:
            all_names = [sk.name] + getattr(
                sk, "name_history", [])
            for name in all_names:
                if (len(name) > 4
                        and name.lower()
                        in resp_lower):
                    return True, sk.skill_id

        # ── 层4：代码内容推断（最后兜底）────────────
        # 如果 response 有代码但没有明确说 skill
        # 根据代码内容和 skill 描述关键词匹配
        code = self._extract_code(response)
        if code:
            code_lower = code.lower()
            skill_scores: dict[str, int] = {}
            for sk in library:
                score = 0
                # 用 description 关键词匹配代码
                desc_words = re.findall(
                    r'\b\w{4,}\b',
                    sk.description.lower())
                for word in set(desc_words):
                    if word in code_lower:
                        score += 1
                # 用 skill_id 关键词匹配
                id_parts = sk.skill_id.replace(
                    "spreadsheet__spreadsheetbench-",
                    "").split("-")
                for part in id_parts:
                    if part in code_lower:
                        score += 2
                if score > 0:
                    skill_scores[sk.skill_id] = score

            if skill_scores:
                best_sid = max(
                    skill_scores,
                    key=skill_scores.get)
                # 至少2分才认为是有效推断
                if skill_scores[best_sid] >= 2:
                    return True, best_sid

        return False, None

    # ──────────────────────────────────────────────
    # 核心评分
    # ──────────────────────────────────────────────

    def _compute_reward(
        self,
        task:             dict,
        steps:            list[TrajectoryStep],
        activation_mode:  str,
        activated_skills: list[str],
        current_skill_id: str,
    ) -> tuple[float, dict]:
        if not steps:
            return 0.0, {"method": "no_steps"}

        spreadsheet  = task.get("spreadsheet", {})
        init_file    = spreadsheet.get("init_file", "")
        golden_file  = spreadsheet.get("golden_file", "")
        answer_pos   = spreadsheet.get("answer_position", "")
        answer_sheet = spreadsheet.get("answer_sheet", "")

        if (golden_file and os.path.exists(golden_file)
                and init_file and os.path.exists(init_file)
                and answer_pos):

            best_score  = 0.0
            best_detail = {
                "method": "golden_compare",
                "error":  "no_code_generated",
                "score":  0.0,
            }

            for step in steps:
                code = self._extract_code(step.action)
                if not code:
                    continue
                score, detail = self._execute_and_score(
                    code         = code,
                    init_file    = init_file,
                    golden_file  = golden_file,
                    answer_pos   = answer_pos,
                    answer_sheet = answer_sheet,
                )
                if score > best_score:
                    best_score  = score
                    best_detail = detail

            best_detail["method"]          = "golden_compare"
            best_detail["skill_activated"] = bool(activated_skills)
            best_detail["skill_id"]        = current_skill_id or "0"
            return best_score, best_detail

        # 正式 reward estimator 必须来自 verifier/golden artifact。
        # 缺失时 fail closed，不以文本长度或关键词 proxy 冒充 Y。
        return 0.0, {
            "method": "invalid_artifact",
            "reason": "missing_init_golden_or_answer_range",
            "valid_for_estimation": False,
        }

    # ──────────────────────────────────────────────
    # 自动补全 save
    # ──────────────────────────────────────────────

    def _ensure_save(self, code: str) -> str:
        save_patterns = [
            "wb.save(", "workbook.save(",
            ".save(wb_path)",
        ]
        if any(p in code for p in save_patterns):
            return code
        if "wb = " in code or "wb=" in code:
            return (code.rstrip()
                    + "\n\n# Auto-save\n"
                    "wb.save(wb_path)\n")
        elif "workbook = " in code:
            return (code.rstrip()
                    + "\n\n# Auto-save\n"
                    "workbook.save(wb_path)\n")
        else:
            return (code.rstrip()
                    + "\n\n# Auto-save\ntry:\n"
                    "    wb.save(wb_path)\n"
                    "except NameError:\n"
                    "    pass\n")

    # ──────────────────────────────────────────────
    # 提取代码
    # ──────────────────────────────────────────────

    def _extract_code(self, response: str) -> str:
        if not response:
            return ""

        # ── 策略 1：fenced code block（多种变体）──
        # 包括 ```python, ```Python, ```py, ```
        # 以及无换行的 ```python code``` 单行形式
        patterns = [
            r"```python\s*\n?(.*?)```",
            r"```Python\s*\n?(.*?)```",
            r"```py\s*\n?(.*?)```",
            r"```\s*\n(.*?)```",
        ]
        for pattern in patterns:
            for m in re.finditer(
                pattern, response,
                re.DOTALL | re.IGNORECASE,
            ):
                code = m.group(1).strip()
                if (len(code) > 10 and
                        self._looks_like_python(code)):
                    return code

        # ── 策略 2：XML-style 标签（Sonnet 常用）──
        for tag in ["code", "python", "solution"]:
            m = re.search(
                rf"<{tag}[^>]*>(.*?)</{tag}>",
                response, re.DOTALL | re.IGNORECASE)
            if m:
                code = m.group(1).strip()
                if (len(code) > 10 and
                        self._looks_like_python(code)):
                    return code

        # ── 策略 3：从关键 import 开始截取 ────
        # Sonnet 有时直接输出代码不加 fence
        import_patterns = [
            r"(from openpyxl[^\n]*\n(?:.*\n)*?"
            r".*?wb\.save\([^)]*\))",
            r"(import openpyxl[^\n]*\n(?:.*\n)*?"
            r".*?wb\.save\([^)]*\))",
            r"(from openpyxl.*?wb\.save\("
            r"wb_path\))",
        ]
        for pattern in import_patterns:
            m = re.search(pattern, response,
                          re.DOTALL)
            if m:
                code = m.group(1).strip()
                if len(code) > 30:
                    return code

        # ── 策略 4：整段回退（宽松）──
        if any(kw in response for kw in [
            "import openpyxl", "import pandas",
            "pd.read_excel", "load_workbook",
            "from openpyxl",
        ]):
            # 去掉 markdown 说明性文字
            lines = response.split("\n")
            code_lines = []
            in_prose = True
            for line in lines:
                stripped = line.strip()
                if not stripped:
                    if not in_prose:
                        code_lines.append(line)
                    continue
                if stripped.startswith("```"):
                    continue
                # 遇到 Python 代码行 → 开始收集
                if (stripped.startswith(
                        ("import ", "from ", "wb",
                         "ws", "if ", "for ",
                         "def ", "#", "try:",
                         "except", "with ",
                         "pd.", "openpyxl.")) or
                        "=" in stripped or
                        "(" in stripped):
                    in_prose = False
                    code_lines.append(line)
                elif not in_prose:
                    code_lines.append(line)
            code = "\n".join(code_lines).strip()
            if len(code) > 30:
                return code

        return ""

    @staticmethod
    def _looks_like_python(code: str) -> bool:
        """快速判断：是否像 Python 代码块。"""
        py_markers = [
            "import ", "from ", "def ", "wb",
            "load_workbook", "openpyxl", "pandas",
            "pd.", "=", "for ", "if ",
        ]
        return any(m in code for m in py_markers)

    # ──────────────────────────────────────────────
    # 执行代码并评分（线程安全缓存）
    # ──────────────────────────────────────────────

    def _execute_and_score(
        self,
        code:         str,
        init_file:    str,
        golden_file:  str,
        answer_pos:   str,
        answer_sheet: str,
    ) -> tuple[float, dict]:
        cache_key = hashlib.md5(
              (code + init_file + golden_file
             + answer_pos + answer_sheet
               + self._file_fingerprint(init_file)
               + self._file_fingerprint(golden_file)
             ).encode()
        ).hexdigest()

        with self._cache_lock:
            if cache_key in self._exec_cache:
                return self._exec_cache[cache_key]

        result = self._do_execute(
            code, init_file, golden_file,
            answer_pos, answer_sheet)

        with self._cache_lock:
            self._exec_cache[cache_key] = result
        return result

    @staticmethod
    def _file_fingerprint(path: str) -> str:
        try:
            stat = os.stat(path)
            return f"{path}:{stat.st_size}:{stat.st_mtime_ns}"
        except OSError:
            return f"{path}:missing"

    def _do_execute(
        self,
        code:         str,
        init_file:    str,
        golden_file:  str,
        answer_pos:   str,
        answer_sheet: str,
    ) -> tuple[float, dict]:
        # 信号量限制并发 subprocess，避免 fork bomb
        # 每个 subprocess 至多 30s，最多 4 个同时
        with self._exec_semaphore, \
                tempfile.TemporaryDirectory() as tmpdir:
            tmp_wb = os.path.join(
                tmpdir, "workbook.xlsx")
            shutil.copy(init_file, tmp_wb)

            exec_script = (
                "import sys, os, warnings\n"
                "warnings.filterwarnings("
                "'ignore', category=UserWarning, "
                "module='openpyxl')\n"
                "try:\n"
                "    import openpyxl\n"
                "    import pandas as pd\n"
                "    from openpyxl import load_workbook\n"
                "    from openpyxl.utils import "
                "get_column_letter, "
                "column_index_from_string\n"
                "except ImportError as e:\n"
                "    print(f'IMPORT_ERROR: {e}',"
                " file=sys.stderr)\n"
                "    sys.exit(1)\n\n"
                f"wb_path = r'{tmp_wb}'\n\n"
                "try:\n"
                + self._indent(code, 4)
                + "\nexcept Exception as e:\n"
                "    print(f'EXEC_ERROR: {e}',"
                " file=sys.stderr)\n"
                "    sys.exit(2)\n\n"
                "print('SUCCESS')\n"
            )

            script_path = os.path.join(
                tmpdir, "solution.py")
            with open(script_path, "w",
                      encoding="utf-8") as f:
                f.write(exec_script)

            # 关键：去掉 preexec_fn（多线程 fork bomb 元凶）
            # 用 Popen + 显式 kill，避免 subprocess.run 卡死
            # 多线程环境下 subprocess.run(timeout=...) 可能
            # 因为 fork 死锁而永远不返回
            popen_kwargs = {
                "stdout": subprocess.PIPE,
                "stderr": subprocess.PIPE,
                "cwd":    tmpdir,
                "text":   True,
                # 不向模型生成代码暴露 API keys/凭据；仅保留最小运行环境。
                "env": {
                    "PATH": os.environ.get("PATH", ""),
                    "PYTHONIOENCODING": "utf-8",
                    "PYTHONDONTWRITEBYTECODE": "1",
                },
            }
            # start_new_session 让子进程有独立 pgid，
            # 我们可以 kill 整个进程组
            if platform.system() != "Windows":
                popen_kwargs["start_new_session"] = True

            proc = None
            try:
                proc = subprocess.Popen(
                    ["python", script_path],
                    **popen_kwargs,
                )
                try:
                    # Excel 操作正常 <3s；给 8s 上限（更快失败）
                    stdout, stderr = proc.communicate(
                        timeout=8)
                    returncode = proc.returncode
                except subprocess.TimeoutExpired:
                    # 强制 kill 整个进程组，快速返回
                    self._kill_process_tree(proc)
                    try:
                        stdout, stderr = (
                            proc.communicate(
                                timeout=1))
                    except Exception:
                        stdout, stderr = "", ""
                    return 0.0, {
                        "error":   "timeout_8s",
                        "matched": 0, "total": 0}
            except Exception as e:
                if proc is not None:
                    try:
                        self._kill_process_tree(proc)
                    except Exception:
                        pass
                return 0.0, {
                    "error":   str(e),
                    "matched": 0, "total": 0}

            if returncode != 0:
                return 0.0, {
                    "error":   (stderr or "")[:300],
                    "matched": 0, "total": 0}

            score, matched, total = \
                self._compare_excel(
                    result_file  = tmp_wb,
                    golden_file  = golden_file,
                    answer_pos   = answer_pos,
                    answer_sheet = answer_sheet,
                )
            return score, {
                "score":   score,
                "matched": matched,
                "total":   total}

    def _indent(self, code: str, n: int) -> str:
        indent = " " * n
        return "\n".join(
            indent + l for l in code.split("\n"))

    @staticmethod
    def _kill_process_tree(proc):
        """
        强制杀掉子进程及其后代（防止孤儿）。
        Linux: 通过 pgid 一次杀整组
        Windows: taskkill /T /F
        """
        try:
            if platform.system() == "Windows":
                subprocess.run(
                    ["taskkill", "/F", "/T",
                     "/PID", str(proc.pid)],
                    capture_output=True,
                    timeout=3,
                )
            else:
                import signal as _sig
                try:
                    os.killpg(
                        os.getpgid(proc.pid),
                        _sig.SIGKILL)
                except (ProcessLookupError,
                        PermissionError):
                    proc.kill()
        except Exception:
            try:
                proc.kill()
            except Exception:
                pass

    # ──────────────────────────────────────────────
    # 对比 Excel
    # ──────────────────────────────────────────────

    def _compare_excel(
        self,
        result_file:  str,
        golden_file:  str,
        answer_pos:   str,
        answer_sheet: str,
    ) -> tuple[float, int, int]:
        try:
            import warnings
            import openpyxl
            warnings.filterwarnings(
                "ignore",
                category=UserWarning,
                module="openpyxl")
            sheet_name = (
                answer_sheet.split(",")[0]
                .strip().strip("'\""))
            pos = (answer_pos.split(",")[0]
                   .strip().strip("'\""))
            if "!" in pos:
                pos = pos.split("!")[-1].strip("'\"")

            wb_r = openpyxl.load_workbook(
                result_file, data_only=True)
            wb_g = openpyxl.load_workbook(
                golden_file, data_only=True)

            ws_r = (wb_r[sheet_name]
                    if sheet_name in wb_r.sheetnames
                    else wb_r.active)
            ws_g = (wb_g[sheet_name]
                    if sheet_name in wb_g.sheetnames
                    else wb_g.active)

            try:
                rv = [[self._norm(c.value)
                        for c in row]
                       for row in ws_r[pos]]
                gv = [[self._norm(c.value)
                        for c in row]
                       for row in ws_g[pos]]
            except Exception:
                return 0.0, 0, 0

            total = matched = 0
            for r_row, g_row in zip(rv, gv):
                for r, g in zip(r_row, g_row):
                    total += 1
                    if self._match(r, g):
                        matched += 1

            if total == 0:
                return 0.0, 0, 0
            return matched / total, matched, total
        except Exception:
            return 0.0, 0, 0

    def _norm(self, v) -> str:
        if v is None:
            return ""
        if isinstance(v, float):
            if v == int(v):
                return str(int(v))
            return (f"{v:.6f}".rstrip("0")
                    .rstrip("."))
        if isinstance(v, int):
            return str(v)
        return str(v).strip()

    def _match(self, a: str, b: str) -> bool:
        if a == b:
            return True
        try:
            fa, fb = float(a), float(b)
            if fb == 0:
                return fa == 0
            return abs(fa - fb) / abs(fb) < 0.01
        except (ValueError, TypeError):
            pass
        return a.lower() == b.lower()

    # ──────────────────────────────────────────────
    # 启发式评分（fallback）
    # ──────────────────────────────────────────────

    def _heuristic_score(
        self,
        last_action:      str,
        activation_mode:  str,
        activated_skills: list[str],
        task:             dict,
    ) -> tuple[float, dict]:
        if activation_mode == "implicit":
            expected = task.get(
                "expected_skill_id", "")
            if not activated_skills:
                return 0.0, {
                    "method": "heuristic_dispatch",
                    "reason": "no_skill_activated"}
            if expected in activated_skills:
                return 1.0, {
                    "method": "heuristic_dispatch",
                    "reason": "correct_skill"}
            return 0.2, {
                "method": "heuristic_dispatch",
                "reason": "wrong_skill"}

        has_code  = "```" in last_action
        has_logic = any(kw in last_action for kw in [
            "openpyxl", "pandas", "pd.",
            "df.", "load_workbook", "wb.save"])
        length = len(last_action)

        if has_code and has_logic and length > 300:
            score = 0.8
        elif has_code and length > 200:
            score = 0.6
        elif length > 300:
            score = 0.4
        elif length > 100:
            score = 0.2
        else:
            score = 0.1

        return score, {
            "method":   "heuristic_quality",
            "has_code": has_code,
            "length":   length}

    def clear_cache(self):
        with self._cache_lock:
            self._exec_cache.clear()

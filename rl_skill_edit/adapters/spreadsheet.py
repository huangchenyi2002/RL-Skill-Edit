"""Forced-skill SpreadsheetBench Student runtime."""

from __future__ import annotations

import math
import os
import posixpath
import signal
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Mapping
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
from openpyxl.utils.exceptions import InvalidFileException

from ..types import SkillArtifact


_EXCEL_MAX_COLUMN = 16_384
_EXCEL_MAX_ROW = 1_048_576
_OFFICE_DOCUMENT_RELATIONSHIPS_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)


@dataclass(frozen=True)
class ExecutionResult:
    score: float
    matched: int
    total: int
    error: str = ""


@dataclass(frozen=True)
class StudentTrajectory:
    task_id: str
    hard_reward: float
    soft_reward: float
    final_answer: str
    visible_logs: tuple[str, ...]
    total_tokens: int
    total_cost_usd: float
    evaluation_valid: bool = True
    invalid_reason: str = ""


@dataclass(frozen=True)
class _WorksheetCellMetadata:
    empty_numeric_cells: frozenset[tuple[int, int]]
    empty_string_formula_cells: frozenset[tuple[int, int]]


def _load_workbook_views(path: Path) -> tuple[Any, Any]:
    raw_workbook = load_workbook(path, data_only=False)
    try:
        cached_workbook = load_workbook(path, data_only=True)
    except BaseException:
        raw_workbook.close()
        raise
    return raw_workbook, cached_workbook


def _merge_topology(
    worksheet: Any,
    *,
    min_column: int,
    min_row: int,
    max_column: int,
    max_row: int,
) -> tuple[tuple[int, int, int, int], ...]:
    intersections = []
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.max_col < min_column
            or merged_range.min_col > max_column
            or merged_range.max_row < min_row
            or merged_range.min_row > max_row
        ):
            continue
        intersections.append(
            (
                merged_range.min_col,
                merged_range.min_row,
                merged_range.max_col,
                merged_range.max_row,
            )
        )
    return tuple(sorted(intersections))


def _strict_cell_value(
    raw_worksheet: Any,
    cached_worksheet: Any,
    *,
    row: int,
    column: int,
    metadata: _WorksheetCellMetadata,
) -> tuple[tuple[str, Any] | None, str]:
    raw_cell = raw_worksheet.cell(row=row, column=column)
    if isinstance(raw_cell, MergedCell):
        merged_range = next(
            (
                candidate
                for candidate in raw_worksheet.merged_cells.ranges
                if candidate.min_row <= row <= candidate.max_row
                and candidate.min_col <= column <= candidate.max_col
            ),
            None,
        )
        if merged_range is None:
            return None, "unsupported_cell_value"
        row = merged_range.min_row
        column = merged_range.min_col
        raw_cell = raw_worksheet.cell(row=row, column=column)
    cached_cell = cached_worksheet.cell(row=row, column=column)
    if raw_cell.data_type == "e" or cached_cell.data_type == "e":
        return None, "unsupported_cell_value"
    if raw_cell.data_type == "f":
        if (row, column) in metadata.empty_string_formula_cells:
            value = ""
        elif cached_cell.value is None:
            return None, "missing_formula_cache"
        else:
            value = cached_cell.value
    elif (row, column) in metadata.empty_numeric_cells:
        return None, "unsupported_cell_value"
    elif raw_cell.data_type == "inlineStr" and raw_cell.value is None:
        value = ""
    else:
        value = cached_cell.value

    if value is None:
        return ("blank", None), ""
    if type(value) is bool:
        return ("boolean", value), ""
    if type(value) in {int, float}:
        if not math.isfinite(float(value)):
            return None, "unsupported_cell_value"
        return ("number", value), ""
    if type(value) is str:
        return ("string", value), ""
    if type(value) in {date, datetime, time, timedelta}:
        return (f"temporal:{type(value).__name__}", value), ""
    return None, "unsupported_cell_value"


def _xml_root_namespace(root: ElementTree.Element, expected_name: str) -> str:
    tag = root.tag
    if not isinstance(tag, str) or not tag.startswith("{"):
        raise ValueError(f"{expected_name} root must have a namespace")
    namespace, separator, name = tag[1:].partition("}")
    if not separator or not namespace or name != expected_name:
        raise ValueError(f"invalid {expected_name} root")
    return namespace


def _read_unique_archive_member(archive: ZipFile, member_name: str) -> bytes:
    matches = tuple(info for info in archive.infolist() if info.filename == member_name)
    if len(matches) != 1:
        raise ValueError(f"archive member is not unique: {member_name}")
    return archive.read(matches[0])


def _normalize_workbook_target(target: str) -> str:
    if "\\" in target or target.startswith("//"):
        raise ValueError("invalid workbook relationship target")
    archive_path = (
        posixpath.normpath(target[1:])
        if target.startswith("/")
        else posixpath.normpath(posixpath.join("xl", target))
    )
    if archive_path in {"", ".", ".."} or not archive_path.startswith("xl/"):
        raise ValueError("workbook relationship target escapes archive")
    return archive_path


def _worksheet_archive_path(archive: ZipFile, sheet_name: str) -> str:
    workbook_root = ElementTree.fromstring(
        _read_unique_archive_member(archive, "xl/workbook.xml")
    )
    workbook_namespace = _xml_root_namespace(workbook_root, "workbook")
    sheets_tag = f"{{{workbook_namespace}}}sheets"
    sheet_tag = f"{{{workbook_namespace}}}sheet"
    sheets_elements = tuple(
        element for element in workbook_root if element.tag == sheets_tag
    )
    if len(sheets_elements) != 1:
        raise ValueError("workbook must contain exactly one sheets element")
    sheet_elements = tuple(
        element for element in sheets_elements[0] if element.tag == sheet_tag
    )
    sheet_names = tuple(element.attrib.get("name") for element in sheet_elements)
    if any(not name for name in sheet_names) or len(set(sheet_names)) != len(
        sheet_names
    ):
        raise ValueError("worksheet names must be present and unique")
    matching_sheets = tuple(
        element
        for element in sheet_elements
        if element.attrib.get("name") == sheet_name
    )
    if len(matching_sheets) != 1:
        raise ValueError(f"missing worksheet relationship: {sheet_name}")
    relationship_attribute = f"{{{_OFFICE_DOCUMENT_RELATIONSHIPS_NAMESPACE}}}id"
    sheet_relationship_ids = tuple(
        element.attrib.get(relationship_attribute) for element in sheet_elements
    )
    if any(not item_id for item_id in sheet_relationship_ids) or len(
        set(sheet_relationship_ids)
    ) != len(sheet_relationship_ids):
        raise ValueError("worksheet relationship ids must be present and unique")
    relationship_id = matching_sheets[0].attrib[relationship_attribute]

    relationships_root = ElementTree.fromstring(
        _read_unique_archive_member(archive, "xl/_rels/workbook.xml.rels")
    )
    relationships_namespace = _xml_root_namespace(
        relationships_root,
        "Relationships",
    )
    relationship_tag = f"{{{relationships_namespace}}}Relationship"
    relationships = tuple(
        element for element in relationships_root if element.tag == relationship_tag
    )
    relationship_ids = tuple(element.attrib.get("Id") for element in relationships)
    if any(not item_id for item_id in relationship_ids) or len(
        set(relationship_ids)
    ) != len(relationship_ids):
        raise ValueError("workbook relationship ids must be present and unique")
    if any(item_id not in relationship_ids for item_id in sheet_relationship_ids):
        raise ValueError("worksheet relationship is missing its target")
    relationship_targets = tuple(
        element.attrib.get("Target") for element in relationships
    )
    if any(not target for target in relationship_targets):
        raise ValueError("workbook relationship targets must be present")
    target_modes = tuple(element.attrib.get("TargetMode") for element in relationships)
    if any(mode not in {None, "Internal", "External"} for mode in target_modes):
        raise ValueError("invalid workbook relationship target mode")
    normalized_targets = tuple(
        _normalize_workbook_target(target)
        for mode, target in zip(
            target_modes,
            relationship_targets,
            strict=True,
        )
        if mode in {None, "Internal"}
    )
    if len(set(normalized_targets)) != len(normalized_targets):
        raise ValueError("internal workbook relationship targets must be unique")
    matching_relationships = tuple(
        element
        for element in relationships
        if element.attrib.get("Id") == relationship_id
    )
    if len(matching_relationships) != 1:
        raise ValueError(f"missing worksheet target: {sheet_name}")
    selected_relationship = matching_relationships[0]
    target = selected_relationship.attrib.get("Target")
    if not target:
        raise ValueError(f"missing worksheet target: {sheet_name}")
    if selected_relationship.attrib.get("TargetMode") not in {None, "Internal"}:
        raise ValueError("worksheet target must be internal")
    archive_path = _normalize_workbook_target(target)
    _read_unique_archive_member(archive, archive_path)
    return archive_path


def _worksheet_cell_metadata(
    workbook_path: Path,
    sheet_name: str,
) -> _WorksheetCellMetadata:
    with ZipFile(workbook_path) as archive:
        worksheet_path = _worksheet_archive_path(archive, sheet_name)
        worksheet_root = ElementTree.fromstring(
            _read_unique_archive_member(archive, worksheet_path)
        )
    worksheet_namespace = _xml_root_namespace(worksheet_root, "worksheet")
    sheet_data_tag = f"{{{worksheet_namespace}}}sheetData"
    row_tag = f"{{{worksheet_namespace}}}row"
    cell_tag = f"{{{worksheet_namespace}}}c"
    formula_tag = f"{{{worksheet_namespace}}}f"
    value_tag = f"{{{worksheet_namespace}}}v"
    sheet_data_elements = tuple(
        element for element in worksheet_root if element.tag == sheet_data_tag
    )
    if len(sheet_data_elements) != 1:
        raise ValueError("worksheet must contain exactly one sheetData element")
    empty_numeric_cells = set()
    empty_string_formula_cells = set()
    seen_coordinates = set()
    row_number = 0
    for row_element in sheet_data_elements[0]:
        if row_element.tag != row_tag:
            continue
        row_reference = row_element.attrib.get("r")
        row_number = int(row_reference) if row_reference is not None else row_number + 1
        column_number = 0
        for element in row_element:
            if element.tag != cell_tag:
                continue
            coordinate = element.attrib.get("r")
            if coordinate is None:
                column_number += 1
                cell_coordinate = (row_number, column_number)
            else:
                cell_coordinate = coordinate_to_tuple(coordinate)
                column_number = cell_coordinate[1]
            if cell_coordinate in seen_coordinates:
                raise ValueError(f"duplicate cell coordinate: {cell_coordinate}")
            seen_coordinates.add(cell_coordinate)

            children = tuple(element)
            formulas = tuple(child for child in children if child.tag == formula_tag)
            values = tuple(child for child in children if child.tag == value_tag)
            if len(formulas) > 1 or len(values) > 1:
                raise ValueError(f"ambiguous cell metadata: {cell_coordinate}")
            if formulas:
                if (
                    element.attrib.get("t") == "str"
                    and len(values) == 1
                    and values[0].text in {None, ""}
                ):
                    empty_string_formula_cells.add(cell_coordinate)
                continue
            if element.attrib.get("t") not in {None, "n"}:
                continue
            if not values or all((value.text or "").strip() for value in values):
                continue
            empty_numeric_cells.add(cell_coordinate)
    return _WorksheetCellMetadata(
        empty_numeric_cells=frozenset(empty_numeric_cells),
        empty_string_formula_cells=frozenset(empty_string_formula_cells),
    )


def _answer_range_coordinates(
    answer_position: str,
) -> tuple[int, int, int, int] | None:
    try:
        coordinates = range_boundaries(answer_position)
    except (TypeError, ValueError):
        return None
    if any(type(value) is not int for value in coordinates):
        return None
    min_column, min_row, max_column, max_row = coordinates
    if (
        min_column < 1
        or min_row < 1
        or max_column < min_column
        or max_row < min_row
        or max_column > _EXCEL_MAX_COLUMN
        or max_row > _EXCEL_MAX_ROW
    ):
        return None
    return min_column, min_row, max_column, max_row


class SpreadsheetExecutor:
    _TIMEOUT_SECONDS = 8

    def execute_and_score(
        self,
        *,
        code: str,
        init_file: str,
        golden_file: str,
        answer_position: str,
        answer_sheet: str,
    ) -> ExecutionResult:
        if not isinstance(code, str) or not code.strip():
            return ExecutionResult(0.0, 0, 0, "missing_executable_code")
        input_path = Path(init_file)
        if not input_path.is_file():
            return ExecutionResult(0.0, 0, 0, "missing_input_workbook")
        try:
            input_workbook = load_workbook(
                input_path,
                data_only=True,
                read_only=True,
            )
        except (BadZipFile, InvalidFileException, OSError, ValueError):
            return ExecutionResult(0.0, 0, 0, "invalid_input_workbook")
        input_workbook.close()
        golden_path = Path(golden_file)
        if not golden_path.is_file():
            return ExecutionResult(0.0, 0, 0, "missing_golden_workbook")
        if not isinstance(answer_position, str) or not answer_position.strip():
            return ExecutionResult(0.0, 0, 0, "missing_answer_position")
        if not isinstance(answer_sheet, str) or not answer_sheet.strip():
            return ExecutionResult(0.0, 0, 0, "missing_answer_sheet")
        if _answer_range_coordinates(answer_position) is None:
            return ExecutionResult(0.0, 0, 0, "invalid_answer_position")
        try:
            golden_workbook = load_workbook(
                golden_path,
                data_only=True,
                read_only=True,
            )
        except (BadZipFile, InvalidFileException, OSError, ValueError):
            return ExecutionResult(0.0, 0, 0, "invalid_golden_workbook")
        try:
            if answer_sheet not in golden_workbook.sheetnames:
                return ExecutionResult(0.0, 0, 0, "missing_golden_sheet")
        finally:
            golden_workbook.close()
        with tempfile.TemporaryDirectory() as temporary_directory:
            temporary_path = Path(temporary_directory)
            workbook_path = temporary_path / "workbook.xlsx"
            shutil.copy2(input_path, workbook_path)
            student_code_path = temporary_path / "student_code.py"
            student_code_path.write_text(code, encoding="utf-8")
            wrapper_path = temporary_path / "runner.py"
            wrapper_path.write_text(
                "import errno\n"
                "import os\n"
                "import sys\n"
                "def _process_isolation_unavailable():\n"
                "    raise RuntimeError('process_isolation_unavailable') from None\n"
                "def _verify_process_isolation():\n"
                "    if os.name == 'nt':\n"
                "        return\n"
                "    import resource\n"
                "    required_os_functions = ('getuid', 'geteuid', 'fork', "
                "'waitpid')\n"
                "    if any(not callable(getattr(os, name, None)) "
                "for name in required_os_functions):\n"
                "        _process_isolation_unavailable()\n"
                "    if not callable(getattr(resource, 'setrlimit', None)) "
                "or not callable(getattr(resource, 'getrlimit', None)):\n"
                "        _process_isolation_unavailable()\n"
                "    try:\n"
                "        real_uid = os.getuid()\n"
                "        effective_uid = os.geteuid()\n"
                "    except BaseException:\n"
                "        _process_isolation_unavailable()\n"
                "    if (type(real_uid) is not int or "
                "type(effective_uid) is not int "
                "or real_uid == 0 or effective_uid == 0):\n"
                "        _process_isolation_unavailable()\n"
                "    if sys.platform == 'linux':\n"
                "        try:\n"
                "            with open('/proc/self/status', "
                "encoding='ascii') as handle:\n"
                "                cap_eff_values = tuple(\n"
                "                    line.split(':', 1)[1].strip()\n"
                "                    for line in handle\n"
                "                    if line.startswith('CapEff:')\n"
                "                )\n"
                "        except BaseException:\n"
                "            _process_isolation_unavailable()\n"
                "        if len(cap_eff_values) != 1:\n"
                "            _process_isolation_unavailable()\n"
                "        try:\n"
                "            effective_capabilities = int(cap_eff_values[0], 16)\n"
                "        except (TypeError, ValueError):\n"
                "            _process_isolation_unavailable()\n"
                "        privileged_capabilities = (1 << 21) | (1 << 24)\n"
                "        if effective_capabilities & privileged_capabilities:\n"
                "            _process_isolation_unavailable()\n"
                "    try:\n"
                "        resource.setrlimit(resource.RLIMIT_NPROC, (0, 0))\n"
                "        process_limits = resource.getrlimit(resource.RLIMIT_NPROC)\n"
                "    except BaseException:\n"
                "        _process_isolation_unavailable()\n"
                "    if (type(process_limits) is not tuple "
                "or len(process_limits) != 2 "
                "or any(type(value) is not int or value != 0 "
                "for value in process_limits)):\n"
                "        _process_isolation_unavailable()\n"
                "    try:\n"
                "        probe_pid = os.fork()\n"
                "    except OSError as exc:\n"
                "        if exc.errno != errno.EAGAIN:\n"
                "            _process_isolation_unavailable()\n"
                "    except BaseException:\n"
                "        _process_isolation_unavailable()\n"
                "    else:\n"
                "        if probe_pid == 0:\n"
                "            os._exit(0)\n"
                "        try:\n"
                "            waited_pid, _ = os.waitpid(probe_pid, 0)\n"
                "        except BaseException:\n"
                "            _process_isolation_unavailable()\n"
                "        if waited_pid != probe_pid:\n"
                "            _process_isolation_unavailable()\n"
                "        _process_isolation_unavailable()\n"
                "_verify_process_isolation()\n"
                f"code_path = {str(student_code_path)!r}\n"
                f"namespace = {{'wb_path': {str(workbook_path)!r}, "
                "'__file__': code_path, '__name__': '__main__'}\n"
                "with open(code_path, encoding='utf-8') as handle:\n"
                "    source = handle.read()\n"
                "compiled = compile(source, code_path, 'exec', dont_inherit=True)\n"
                "exec(compiled, namespace, namespace)\n",
                encoding="utf-8",
            )
            process: subprocess.Popen[str] | None = None
            try:
                process = subprocess.Popen(
                    [sys.executable, str(wrapper_path)],
                    cwd=temporary_path,
                    env={
                        "PATH": os.environ.get("PATH", ""),
                        "PYTHONIOENCODING": "utf-8",
                        "PYTHONDONTWRITEBYTECODE": "1",
                    },
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    start_new_session=os.name != "nt",
                )
                try:
                    stdout, stderr = process.communicate(timeout=self._TIMEOUT_SECONDS)
                except subprocess.TimeoutExpired:
                    self._kill_process_tree(process)
                    try:
                        process.communicate(timeout=1)
                    except (OSError, subprocess.SubprocessError):
                        pass
                    return ExecutionResult(0.0, 0, 0, "timeout_8s")
            except OSError as exc:
                if process is not None:
                    self._kill_process_tree(process)
                return ExecutionResult(0.0, 0, 0, f"execution_error:{exc}")
            if process.returncode != 0:
                return ExecutionResult(
                    0.0,
                    0,
                    0,
                    (stderr or stdout or "execution_failed")[:500],
                )
            return self._compare_workbooks(
                workbook_path,
                golden_path,
                answer_position=answer_position,
                answer_sheet=answer_sheet,
            )

    @staticmethod
    def _kill_process_tree(process: subprocess.Popen[str]) -> None:
        try:
            if os.name == "nt":
                subprocess.run(
                    [
                        "taskkill",
                        "/F",
                        "/T",
                        "/PID",
                        str(process.pid),
                    ],
                    capture_output=True,
                    timeout=3,
                    check=False,
                )
            else:
                os.killpg(process.pid, signal.SIGKILL)
        except (OSError, subprocess.SubprocessError):
            try:
                process.kill()
            except OSError:
                pass

    @staticmethod
    def _compare_workbooks(
        result_file: Path,
        golden_file: Path,
        *,
        answer_position: str,
        answer_sheet: str,
    ) -> ExecutionResult:
        if not result_file.is_file():
            return ExecutionResult(0.0, 0, 0, "missing_result_workbook")
        try:
            result_raw, result_cached = _load_workbook_views(result_file)
        except (BadZipFile, InvalidFileException, OSError, ValueError):
            return ExecutionResult(0.0, 0, 0, "invalid_result_workbook")
        try:
            golden_raw, golden_cached = _load_workbook_views(golden_file)
        except (BadZipFile, InvalidFileException, OSError, ValueError):
            result_raw.close()
            result_cached.close()
            return ExecutionResult(0.0, 0, 0, "invalid_golden_workbook")
        try:
            if answer_sheet not in result_raw.sheetnames:
                return ExecutionResult(0.0, 0, 0, "missing_result_sheet")
            if answer_sheet not in golden_raw.sheetnames:
                return ExecutionResult(0.0, 0, 0, "missing_golden_sheet")
            coordinates = _answer_range_coordinates(answer_position)
            if coordinates is None:
                return ExecutionResult(0.0, 0, 0, "invalid_answer_position")
            min_column, min_row, max_column, max_row = coordinates
            result_raw_sheet = result_raw[answer_sheet]
            result_cached_sheet = result_cached[answer_sheet]
            golden_raw_sheet = golden_raw[answer_sheet]
            golden_cached_sheet = golden_cached[answer_sheet]
            try:
                result_metadata = _worksheet_cell_metadata(
                    result_file,
                    answer_sheet,
                )
            except (BadZipFile, KeyError, OSError, ValueError, ElementTree.ParseError):
                return ExecutionResult(0.0, 0, 0, "invalid_result_workbook")
            try:
                golden_metadata = _worksheet_cell_metadata(
                    golden_file,
                    answer_sheet,
                )
            except (BadZipFile, KeyError, OSError, ValueError, ElementTree.ParseError):
                return ExecutionResult(0.0, 0, 0, "invalid_golden_workbook")
            range_coordinates = {
                "min_column": min_column,
                "min_row": min_row,
                "max_column": max_column,
                "max_row": max_row,
            }
            if _merge_topology(
                result_raw_sheet,
                **range_coordinates,
            ) != _merge_topology(
                golden_raw_sheet,
                **range_coordinates,
            ):
                return ExecutionResult(0.0, 0, 0, "merge_topology_mismatch")
            matched = 0
            total = 0
            for row in range(min_row, max_row + 1):
                for column in range(min_column, max_column + 1):
                    result_value, result_error = _strict_cell_value(
                        result_raw_sheet,
                        result_cached_sheet,
                        row=row,
                        column=column,
                        metadata=result_metadata,
                    )
                    if result_error:
                        return ExecutionResult(0.0, 0, 0, result_error)
                    golden_value, golden_error = _strict_cell_value(
                        golden_raw_sheet,
                        golden_cached_sheet,
                        row=row,
                        column=column,
                        metadata=golden_metadata,
                    )
                    if golden_error:
                        return ExecutionResult(0.0, 0, 0, golden_error)
                    total += 1
                    if result_value == golden_value:
                        matched += 1
            return ExecutionResult(matched / total, matched, total)
        finally:
            result_raw.close()
            result_cached.close()
            golden_raw.close()
            golden_cached.close()


def build_student_system(
    task: Mapping[str, Any], *, expose_answer_metadata: bool
) -> str:
    metadata = ""
    if expose_answer_metadata:
        spreadsheet = task.get("spreadsheet")
        if isinstance(spreadsheet, Mapping):
            answer_sheet = str(spreadsheet.get("answer_sheet", ""))
            answer_position = str(spreadsheet.get("answer_position", ""))
            if answer_sheet or answer_position:
                metadata = (
                    f"\nTarget sheet: {answer_sheet}\nTarget range: {answer_position}"
                )
    return (
        "You are an AI agent solving an Excel task. Write Python code that "
        "uses the provided wb_path, modifies the workbook, and explicitly "
        "saves it back to wb_path. Return executable Python code."
        f"{metadata}"
    )


def _extract_python(response: str) -> str:
    in_fence = False
    capture = False
    captured_lines: list[str] = []

    for line in response.splitlines():
        stripped = line.strip()
        if not in_fence:
            if not stripped.startswith("```"):
                continue
            language = stripped[3:].strip().lower()
            in_fence = True
            capture = language in {"", "python", "py"}
            captured_lines = []
            continue

        if stripped == "```":
            if capture:
                code = "\n".join(captured_lines).strip()
                if code:
                    return code
            in_fence = False
            capture = False
            captured_lines = []
            continue

        if capture:
            captured_lines.append(line)

    return ""


def _validate_task(
    task: Mapping[str, Any],
) -> tuple[str, str, Mapping[str, Any]]:
    if not isinstance(task, Mapping):
        raise ValueError("task must be a mapping")
    task_id = task.get("task_id")
    if not isinstance(task_id, str) or not task_id.strip():
        raise ValueError("task.task_id must be a non-empty string")
    description = task.get("description")
    if not isinstance(description, str) or not description.strip():
        raise ValueError("task.description must be a non-empty string")
    spreadsheet = task.get("spreadsheet")
    if not isinstance(spreadsheet, Mapping):
        raise ValueError("task.spreadsheet must be a mapping")
    for key in ("init_file", "golden_file"):
        value = spreadsheet.get(key)
        if not isinstance(value, str) or not value.strip():
            raise ValueError(f"task.spreadsheet.{key} must be a file path")
        if not Path(value).is_file():
            raise ValueError(f"task.spreadsheet.{key} does not exist")
    for key in ("answer_sheet", "answer_position"):
        value = spreadsheet.get(key)
        if not isinstance(value, str) or not value.strip():
            raise ValueError(f"task.spreadsheet.{key} must be a non-empty string")
    return task_id, description, spreadsheet


class SpreadsheetStudent:
    def __init__(self, config: Mapping[str, Any], client: Any) -> None:
        student_config = config["student"]
        self.client = client
        self.model = str(student_config["model"])
        self.temperature = float(student_config["temperature"])
        self.max_tokens = int(student_config["max_tokens"])
        self.max_steps = int(student_config["max_steps"])
        self.executor = SpreadsheetExecutor()

    def run_task(
        self,
        task: Mapping[str, Any],
        skill: SkillArtifact,
        *,
        blind: bool,
        seed: int,
    ) -> StudentTrajectory:
        task_id, description, spreadsheet = _validate_task(task)
        system = build_student_system(task, expose_answer_metadata=not blind)
        system += f"\n\n[ACTIVE SKILL: {skill.name}]\n{skill.body}"
        response, usage = self.client.chat(
            model=self.model,
            messages=[{"role": "user", "content": description}],
            system=system,
            temperature=self.temperature,
            max_tokens=self.max_tokens,
            call_type="student_rollout",
            seed=seed,
        )
        if (
            not isinstance(usage, Mapping)
            or not {
                "ok",
                "total_tokens",
                "cost_usd",
            }
            <= usage.keys()
        ):
            return StudentTrajectory(
                task_id=task_id,
                hard_reward=0.0,
                soft_reward=0.0,
                final_answer="",
                visible_logs=(),
                total_tokens=0,
                total_cost_usd=0.0,
                evaluation_valid=False,
                invalid_reason="api:incomplete_usage",
            )
        total_tokens_value = usage["total_tokens"]
        total_cost_value = usage["cost_usd"]
        if (
            not isinstance(usage["ok"], bool)
            or isinstance(total_tokens_value, bool)
            or not isinstance(total_tokens_value, int)
            or total_tokens_value < 0
            or isinstance(total_cost_value, bool)
            or not isinstance(total_cost_value, (int, float))
            or not math.isfinite(float(total_cost_value))
            or float(total_cost_value) < 0.0
        ):
            return StudentTrajectory(
                task_id=task_id,
                hard_reward=0.0,
                soft_reward=0.0,
                final_answer="",
                visible_logs=(),
                total_tokens=0,
                total_cost_usd=0.0,
                evaluation_valid=False,
                invalid_reason="api:invalid_usage",
            )
        total_tokens = total_tokens_value
        total_cost_usd = float(total_cost_value)
        if not usage["ok"]:
            return StudentTrajectory(
                task_id=task_id,
                hard_reward=0.0,
                soft_reward=0.0,
                final_answer="",
                visible_logs=(),
                total_tokens=total_tokens,
                total_cost_usd=total_cost_usd,
                evaluation_valid=False,
                invalid_reason=f"api:{usage.get('error_kind', 'request_failed')}",
            )
        if not response:
            return StudentTrajectory(
                task_id=task_id,
                hard_reward=0.0,
                soft_reward=0.0,
                final_answer="",
                visible_logs=(),
                total_tokens=total_tokens,
                total_cost_usd=total_cost_usd,
                evaluation_valid=False,
                invalid_reason=(
                    f"api:{usage.get('error_kind', 'empty_response')}"
                    if not usage.get("ok", True)
                    else "empty_response"
                ),
            )

        code = _extract_python(response)
        if not code:
            return StudentTrajectory(
                task_id=task_id,
                hard_reward=0.0,
                soft_reward=0.0,
                final_answer=response,
                visible_logs=(),
                total_tokens=total_tokens,
                total_cost_usd=total_cost_usd,
                evaluation_valid=False,
                invalid_reason="missing_executable_code",
            )

        result = self.executor.execute_and_score(
            code=code,
            init_file=str(spreadsheet["init_file"]),
            golden_file=str(spreadsheet["golden_file"]),
            answer_position=str(spreadsheet["answer_position"]),
            answer_sheet=str(spreadsheet["answer_sheet"]),
        )
        return StudentTrajectory(
            task_id=task_id,
            hard_reward=float(result.score),
            soft_reward=float(result.score),
            final_answer=response,
            visible_logs=((result.error,) if result.error else ()),
            total_tokens=total_tokens,
            total_cost_usd=total_cost_usd,
            evaluation_valid=not bool(result.error),
            invalid_reason=result.error,
        )


__all__ = [
    "ExecutionResult",
    "SpreadsheetExecutor",
    "SpreadsheetStudent",
    "StudentTrajectory",
    "build_student_system",
]

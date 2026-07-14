from __future__ import annotations

import copy
import errno
import importlib.util
import json
import os
import signal
import subprocess
from collections.abc import Callable
from datetime import datetime
from pathlib import Path
from types import MappingProxyType
from xml.etree import ElementTree
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

import rl_skill_edit.evaluation as evaluation
from rl_skill_edit.adapters.spreadsheet import (
    ExecutionResult,
    SpreadsheetExecutor,
    SpreadsheetStudent,
    StudentTrajectory,
)
from rl_skill_edit.cache import JsonFileCache
from rl_skill_edit.types import SkillArtifact, Split


STUDENT_CONFIG = {
    "student": {
        "model": "fake-student",
        "temperature": 0.0,
        "max_tokens": 256,
        "max_steps": 3,
    }
}


class RecordingClient:
    def __init__(
        self,
        response: str | None = None,
        usage: dict | None = None,
    ) -> None:
        self.calls: list[dict] = []
        self.response = response
        self.usage = usage

    def chat(self, **kwargs):
        self.calls.append(copy.deepcopy(kwargs))
        response = (
            self.response
            if self.response is not None
            else (
                "```python\n"
                "from openpyxl import load_workbook\n"
                "wb = load_workbook(wb_path)\n"
                "wb.active['A1'] = 'candidate'\n"
                "wb.save(wb_path)\n"
                "```"
            )
        )
        usage = (
            self.usage
            if self.usage is not None
            else {"total_tokens": 1, "cost_usd": 0.0, "ok": True}
        )
        return response, usage


def _write_workbook(path: Path, value: object) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SECRET_FINAL_SHEET"
    worksheet["A1"] = value
    workbook.save(path)


def _rewrite_archive_xml(
    path: Path,
    member_name: str,
    mutate: Callable[[ElementTree.Element], None],
) -> None:
    with ZipFile(path) as source:
        members = [(info, source.read(info.filename)) for info in source.infolist()]
    rewritten_members = []
    found = False
    for info, data in members:
        if info.filename == member_name:
            root = ElementTree.fromstring(data)
            mutate(root)
            data = ElementTree.tostring(root)
            found = True
        rewritten_members.append((info, data))
    assert found
    rewritten_path = path.with_name(f"{path.stem}.rewritten.xlsx")
    with ZipFile(rewritten_path, "w") as destination:
        for info, data in rewritten_members:
            destination.writestr(info, data)
    rewritten_path.replace(path)


def _set_formula_cache_metadata(
    path: Path,
    *,
    cell_type: str,
    include_value: bool,
    append_duplicate_outside_sheet_data: bool = False,
    append_foreign_value: bool = False,
    worksheet_name: str = "xl/worksheets/sheet1.xml",
) -> None:
    def mutate(root: ElementTree.Element) -> None:
        cell = next(
            element
            for element in root.iter()
            if element.tag.rsplit("}", 1)[-1] == "c" and element.attrib.get("r") == "A1"
        )
        cell.set("t", cell_type)
        value_elements = [
            child for child in cell if child.tag.rsplit("}", 1)[-1] == "v"
        ]
        for value_element in value_elements:
            cell.remove(value_element)
        if include_value:
            namespace = cell.tag.split("}", 1)[0].lstrip("{")
            ElementTree.SubElement(cell, f"{{{namespace}}}v")
        if append_foreign_value:
            ElementTree.SubElement(cell, "{urn:evil}v")
        if append_duplicate_outside_sheet_data:
            namespace = cell.tag.split("}", 1)[0].lstrip("{")
            duplicate = ElementTree.SubElement(
                root,
                f"{{{namespace}}}c",
                {"r": "A1", "t": "str"},
            )
            ElementTree.SubElement(
                duplicate, f"{{{namespace}}}f"
            ).text = 'IF(1=1,"","")'
            ElementTree.SubElement(duplicate, f"{{{namespace}}}v")

    _rewrite_archive_xml(path, worksheet_name, mutate)


def _workbook_relationship_id(path: Path, sheet_name: str) -> str:
    office_relationship_namespace = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    namespace = root.tag.split("}", 1)[0].lstrip("{")
    sheets = next(child for child in root if child.tag == f"{{{namespace}}}sheets")
    sheet = next(
        child
        for child in sheets
        if child.tag == f"{{{namespace}}}sheet"
        and child.attrib.get("name") == sheet_name
    )
    return sheet.attrib[f"{{{office_relationship_namespace}}}id"]


def _append_fake_sheet_outside_sheets(
    path: Path,
    *,
    fake_name: str,
    target_sheet_name: str,
) -> None:
    relationship_id = _workbook_relationship_id(path, target_sheet_name)
    office_relationship_namespace = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )

    def mutate(root: ElementTree.Element) -> None:
        namespace = root.tag.split("}", 1)[0].lstrip("{")
        fake = ElementTree.Element(
            f"{{{namespace}}}sheet",
            {
                "name": fake_name,
                "sheetId": "999",
                f"{{{office_relationship_namespace}}}id": relationship_id,
            },
        )
        root.insert(0, fake)

    _rewrite_archive_xml(path, "xl/workbook.xml", mutate)


def _duplicate_workbook_relationship_id(path: Path, *, sheet_name: str) -> None:
    relationship_id = _workbook_relationship_id(path, sheet_name)

    def mutate(root: ElementTree.Element) -> None:
        namespace = root.tag.split("}", 1)[0].lstrip("{")
        tag = f"{{{namespace}}}Relationship"
        relationship = next(
            child
            for child in root
            if child.tag == tag and child.attrib.get("Id") == relationship_id
        )
        root.append(ElementTree.Element(tag, dict(relationship.attrib)))

    _rewrite_archive_xml(path, "xl/_rels/workbook.xml.rels", mutate)


def _execute_without_editing(
    init_file: Path,
    golden_file: Path,
    *,
    answer_position: str = "A1",
) -> ExecutionResult:
    return SpreadsheetExecutor().execute_and_score(
        code="pass",
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position=answer_position,
        answer_sheet="SECRET_FINAL_SHEET",
    )


def _process_exists(process_id: int) -> bool:
    try:
        os.kill(process_id, 0)
    except ProcessLookupError:
        return False
    return True


def _install_runner_startup_hook(monkeypatch, source: str) -> None:
    real_popen = subprocess.Popen

    def instrumented_popen(command, **kwargs):
        runner_path = Path(command[1])
        runner_path.with_name("resource.py").write_text(
            source,
            encoding="utf-8",
        )
        return real_popen(command, **kwargs)

    monkeypatch.setattr(
        "rl_skill_edit.adapters.spreadsheet.subprocess.Popen",
        instrumented_popen,
    )


def _isolation_resource_source(
    *,
    limits: object = (0, 0),
    fork_errno: int | None = errno.EAGAIN,
    uid: int = 1_000,
    euid: int = 1_000,
    cap_eff: str = "0",
    missing_os_function: str | None = None,
) -> str:
    status = f"Name:\ttest\nCapEff:\t{cap_eff}\n"
    source = (
        "import builtins\n"
        "import io\n"
        "import os\n"
        "import sys\n"
        "RLIMIT_NPROC = 1\n"
        "def setrlimit(*args, **kwargs):\n"
        "    pass\n"
        "def getrlimit(*args, **kwargs):\n"
        f"    return {limits!r}\n"
        f"os.getuid = lambda: {uid!r}\n"
        f"os.geteuid = lambda: {euid!r}\n"
        "sys.platform = 'linux'\n"
        "_real_open = builtins.open\n"
        "def _patched_open(file, *args, **kwargs):\n"
        "    if os.fspath(file) == '/proc/self/status':\n"
        f"        return io.StringIO({status!r})\n"
        "    return _real_open(file, *args, **kwargs)\n"
        "builtins.open = _patched_open\n"
    )
    if fork_errno is not None:
        source += (
            "def _probe_fork():\n"
            f"    raise OSError({fork_errno!r}, 'probe failure')\n"
            "os.fork = _probe_fork\n"
        )
    if missing_os_function is not None:
        source += f"delattr(os, {missing_os_function!r})\n"
    return source


def _final_task(tmp_path: Path) -> dict:
    init_file = tmp_path / "final_input.xlsx"
    golden_file = tmp_path / "final_golden.xlsx"
    _write_workbook(init_file, "input")
    _write_workbook(golden_file, "golden")
    return {
        "task_id": "final-1",
        "description": "Update the workbook.",
        "spreadsheet": {
            "init_file": str(init_file),
            "golden_file": str(golden_file),
            "answer_sheet": "SECRET_FINAL_SHEET",
            "answer_position": "B17:C19",
        },
    }


def _evaluation_task(tmp_path: Path, task_id: str) -> dict:
    init_file = tmp_path / f"{task_id}_input.xlsx"
    golden_file = tmp_path / f"{task_id}_golden.xlsx"
    _write_workbook(init_file, f"input-{task_id}")
    _write_workbook(golden_file, f"golden-{task_id}")
    return {
        "task_id": task_id,
        "description": f"Update workbook {task_id}.",
        "spreadsheet": {
            "init_file": str(init_file),
            "golden_file": str(golden_file),
            "answer_sheet": "SECRET_FINAL_SHEET",
            "answer_position": "A1",
        },
    }


class RecordingStudent:
    model = "frozen-student"
    temperature = 0.0
    max_tokens = 256
    max_steps = 3

    def __init__(self) -> None:
        self.calls: list[dict] = []

    def run_task(self, task, skill, *, blind, seed):
        self.calls.append(
            {
                "task_id": task["task_id"],
                "skill_digest": skill.digest,
                "blind": blind,
                "seed": seed,
            }
        )
        hard = seed / 1000.0
        return StudentTrajectory(
            task_id=task["task_id"],
            hard_reward=hard,
            soft_reward=hard + 0.4,
            final_answer=f"answer-{seed}",
            visible_logs=(f"log-{seed}",),
            total_tokens=3,
            total_cost_usd=0.01,
        )


class IncompleteStudent(RecordingStudent):
    def __init__(self, failure: str) -> None:
        super().__init__()
        self.failure = failure

    def run_task(self, task, skill, *, blind, seed):
        trajectory = super().run_task(task, skill, blind=blind, seed=seed)
        if len(self.calls) != 2:
            return trajectory
        if self.failure == "invalid":
            return StudentTrajectory(
                task_id=trajectory.task_id,
                hard_reward=0.0,
                soft_reward=0.0,
                final_answer=trajectory.final_answer,
                visible_logs=trajectory.visible_logs,
                total_tokens=trajectory.total_tokens,
                total_cost_usd=trajectory.total_cost_usd,
                evaluation_valid=False,
                invalid_reason="missing_executable_code",
            )
        return StudentTrajectory(
            task_id="wrong-task-id",
            hard_reward=trajectory.hard_reward,
            soft_reward=trajectory.soft_reward,
            final_answer=trajectory.final_answer,
            visible_logs=trajectory.visible_logs,
            total_tokens=trajectory.total_tokens,
            total_cost_usd=trajectory.total_cost_usd,
        )


class UsageCountingStudent(RecordingStudent):
    def __init__(self) -> None:
        super().__init__()
        self.client = type(
            "UsageClient",
            (),
            {
                "total_input_tokens": 10,
                "total_output_tokens": 20,
                "total_cost_usd": 1.0,
            },
        )()

    def run_task(self, task, skill, *, blind, seed):
        trajectory = super().run_task(task, skill, blind=blind, seed=seed)
        self.client.total_input_tokens += 2
        self.client.total_output_tokens += 3
        self.client.total_cost_usd += 0.05
        return StudentTrajectory(
            task_id=trajectory.task_id,
            hard_reward=trajectory.hard_reward,
            soft_reward=trajectory.soft_reward,
            final_answer=trajectory.final_answer,
            visible_logs=trajectory.visible_logs,
            total_tokens=5,
            total_cost_usd=0.05,
        )


def test_spreadsheet_adapter_module_exists() -> None:
    assert importlib.util.find_spec("rl_skill_edit.adapters.spreadsheet") is not None


def test_spreadsheet_skill_evaluator_exists() -> None:
    assert callable(getattr(evaluation, "SpreadsheetSkillEvaluator", None))


def test_runtime_has_no_legacy_repository_evaluator_or_src_dependency() -> None:
    assert not hasattr(evaluation, "RepositorySkillEvaluator")
    source = Path(evaluation.__file__).read_text(encoding="utf-8")
    assert "src.evaluator" not in source


@pytest.mark.parametrize(
    ("metric", "weight", "expected"),
    [
        ("hard", 0.5, 0.2),
        ("soft", 0.5, 0.8),
        ("mixed", 0.25, 0.35),
        ("mixed", -3.0, 0.2),
        ("mixed", 3.0, 0.8),
    ],
)
def test_select_score_is_local_and_clamps_mixed_weight(
    metric: str,
    weight: float,
    expected: float,
) -> None:
    assert evaluation.select_score(0.2, 0.8, metric, weight) == pytest.approx(expected)


def test_select_score_rejects_an_unknown_metric() -> None:
    with pytest.raises(ValueError, match="unknown metric"):
        evaluation.select_score(0.2, 0.8, "unknown", 0.5)


def test_spreadsheet_evaluator_preserves_order_and_derives_each_seed(
    tmp_path: Path,
) -> None:
    student = RecordingStudent()
    evaluator = evaluation.SpreadsheetSkillEvaluator(
        student,
        cache=None,
        gate_metric="mixed",
        gate_mixed_weight=0.25,
        success_threshold=0.8,
    )
    tasks = (
        _evaluation_task(tmp_path, "task-b"),
        _evaluation_task(tmp_path, "task-a"),
    )

    batch = evaluator.evaluate(
        SkillArtifact("main", "main", "", "## Rule\nUpdate cells."),
        tasks,
        split=Split.TRAIN,
        seed=100,
        repetitions=2,
        use_cache=False,
        blind=False,
    )

    assert batch.ordered_task_ids == ("task-b", "task-a")
    assert [call["seed"] for call in student.calls] == [100, 101, 102, 103]
    assert [call["task_id"] for call in student.calls] == [
        "task-b",
        "task-b",
        "task-a",
        "task-a",
    ]
    assert batch.results[0].raw_rewards == pytest.approx((0.2, 0.201))
    assert batch.results[1].raw_rewards == pytest.approx((0.202, 0.203))
    assert batch.usage["student_rollouts"] == 4


def test_spreadsheet_evaluator_records_client_input_and_output_usage(
    tmp_path: Path,
) -> None:
    student = UsageCountingStudent()
    evaluator = evaluation.SpreadsheetSkillEvaluator(
        student,
        cache=None,
        gate_metric="hard",
        gate_mixed_weight=0.5,
    )

    batch = evaluator.evaluate(
        SkillArtifact("main", "main", "", "## Rule\nUpdate cells."),
        (_evaluation_task(tmp_path, "task-a"),),
        split=Split.TRAIN,
        seed=10,
        repetitions=2,
        use_cache=False,
        blind=False,
    )

    assert batch.usage["input_tokens"] == 4
    assert batch.usage["output_tokens"] == 6
    assert batch.usage["total_tokens"] == 10
    assert batch.usage["cost_usd"] == pytest.approx(0.1)


@pytest.mark.parametrize("failure", ["invalid", "wrong_task_id"])
def test_spreadsheet_evaluator_rejects_an_incomplete_bundle_immediately(
    tmp_path: Path,
    failure: str,
) -> None:
    student = IncompleteStudent(failure)
    evaluator = evaluation.SpreadsheetSkillEvaluator(
        student,
        cache=None,
        gate_metric="hard",
        gate_mixed_weight=0.5,
    )

    with pytest.raises(RuntimeError, match="incomplete evaluation bundle"):
        evaluator.evaluate(
            SkillArtifact("main", "main", "", "## Rule\nUpdate cells."),
            (_evaluation_task(tmp_path, "task-a"),),
            split=Split.TRAIN,
            seed=10,
            repetitions=3,
            use_cache=False,
            blind=False,
        )

    assert len(student.calls) == 2


@pytest.mark.parametrize(
    ("reward_field", "invalid_value"),
    [
        ("hard_reward", -0.01),
        ("hard_reward", 1.01),
        ("soft_reward", -0.01),
        ("soft_reward", 1.01),
    ],
)
def test_spreadsheet_evaluator_rejects_live_rewards_outside_unit_interval(
    tmp_path: Path,
    reward_field: str,
    invalid_value: float,
) -> None:
    student = RecordingStudent()

    def invalid_run_task(task, skill, *, blind, seed):
        values = {"hard_reward": 0.5, "soft_reward": 0.5}
        values[reward_field] = invalid_value
        return StudentTrajectory(
            task_id=task["task_id"],
            hard_reward=values["hard_reward"],
            soft_reward=values["soft_reward"],
            final_answer="answer",
            visible_logs=(),
            total_tokens=1,
            total_cost_usd=0.0,
        )

    student.run_task = invalid_run_task
    evaluator = evaluation.SpreadsheetSkillEvaluator(
        student,
        cache=None,
        gate_metric="hard",
        gate_mixed_weight=0.5,
    )

    with pytest.raises(RuntimeError, match=f"invalid {reward_field}"):
        evaluator.evaluate(
            SkillArtifact("main", "main", "", "## Rule\nUpdate cells."),
            (_evaluation_task(tmp_path, "task-a"),),
            split=Split.TRAIN,
            seed=10,
            repetitions=1,
            use_cache=False,
            blind=False,
        )


def test_spreadsheet_evaluator_requires_freeze_before_test(
    tmp_path: Path,
) -> None:
    student = RecordingStudent()
    evaluator = evaluation.SpreadsheetSkillEvaluator(
        student,
        cache=JsonFileCache(tmp_path / "test-rollouts.json"),
        gate_metric="hard",
        gate_mixed_weight=0.5,
    )
    kwargs = {
        "skill": SkillArtifact("main", "main", "", "## Rule\nUpdate cells."),
        "tasks": (_evaluation_task(tmp_path, "test-a"),),
        "split": Split.TEST,
        "seed": 10,
        "repetitions": 1,
        "use_cache": False,
        "blind": True,
    }

    with pytest.raises(RuntimeError, match="freeze"):
        evaluator.evaluate(**kwargs)
    with pytest.raises(RuntimeError, match="freeze"):
        evaluator.cache_will_hit(
            **{key: value for key, value in kwargs.items() if key != "use_cache"}
        )
    assert student.calls == []

    evaluator.freeze()
    batch = evaluator.evaluate(**kwargs)
    assert batch.ordered_task_ids == ("test-a",)
    assert student.calls[0]["blind"] is True


def test_spreadsheet_evaluator_cache_binds_task_files_and_student_signature(
    tmp_path: Path,
) -> None:
    cache = JsonFileCache(tmp_path / "rollouts.json")
    student = RecordingStudent()
    evaluator = evaluation.SpreadsheetSkillEvaluator(
        student,
        cache=cache,
        gate_metric="hard",
        gate_mixed_weight=0.5,
    )
    skill = SkillArtifact("main", "main", "", "## Rule\nUpdate cells.")
    task = _evaluation_task(tmp_path, "train-a")
    kwargs = {
        "skill": skill,
        "tasks": (task,),
        "split": Split.TRAIN,
        "seed": 10,
        "repetitions": 1,
        "blind": False,
    }

    assert evaluator.cache_will_hit(**kwargs) is False
    assert evaluator.evaluate(**kwargs, use_cache=True).cache_hit is False
    assert evaluator.cache_will_hit(**kwargs) is True
    assert evaluator.evaluate(**kwargs, use_cache=True).cache_hit is True
    assert len(student.calls) == 1

    changed_task = copy.deepcopy(task)
    changed_task["description"] = "Changed task content with the same task_id."
    assert (
        evaluator.evaluate(
            **{**kwargs, "tasks": (changed_task,)},
            use_cache=True,
        ).cache_hit
        is False
    )

    _write_workbook(Path(task["spreadsheet"]["init_file"]), "changed-input")
    assert evaluator.evaluate(**kwargs, use_cache=True).cache_hit is False

    changed_student = RecordingStudent()
    changed_student.model = "different-frozen-student"
    changed_evaluator = evaluation.SpreadsheetSkillEvaluator(
        changed_student,
        cache=cache,
        gate_metric="hard",
        gate_mixed_weight=0.5,
    )
    assert changed_evaluator.evaluate(**kwargs, use_cache=True).cache_hit is False


@pytest.mark.parametrize("tamper", ["task_order", "repetitions"])
def test_spreadsheet_evaluator_rejects_an_incomplete_cached_bundle(
    tmp_path: Path,
    tamper: str,
) -> None:
    cache_path = tmp_path / "rollouts.json"
    cache = JsonFileCache(cache_path)
    student = RecordingStudent()
    evaluator = evaluation.SpreadsheetSkillEvaluator(
        student,
        cache=cache,
        gate_metric="hard",
        gate_mixed_weight=0.5,
    )
    kwargs = {
        "skill": SkillArtifact("main", "main", "", "## Rule\nUpdate cells."),
        "tasks": (
            _evaluation_task(tmp_path, "task-a"),
            _evaluation_task(tmp_path, "task-b"),
        ),
        "split": Split.TRAIN,
        "seed": 10,
        "repetitions": 2,
        "use_cache": True,
        "blind": False,
    }
    evaluator.evaluate(**kwargs)
    payload = json.loads(cache_path.read_text(encoding="utf-8"))
    cached_batch = next(iter(payload["rollout"].values()))
    if tamper == "task_order":
        cached_batch["results"].reverse()
    else:
        cached_batch["results"][0]["raw_rewards"].pop()
    cache_path.write_text(
        json.dumps(payload),
        encoding="utf-8",
    )
    student.calls.clear()

    with pytest.raises(RuntimeError, match="incomplete cached evaluation bundle"):
        evaluator.evaluate(**kwargs)
    assert student.calls == []


def test_spreadsheet_evaluator_rejects_coerced_cached_result_fields(
    tmp_path: Path,
) -> None:
    cache_path = tmp_path / "rollouts.json"
    student = RecordingStudent()
    evaluator = evaluation.SpreadsheetSkillEvaluator(
        student,
        cache=JsonFileCache(cache_path),
        gate_metric="soft",
        gate_mixed_weight=0.5,
    )
    kwargs = {
        "skill": SkillArtifact("main", "main", "", "## Rule\nUpdate cells."),
        "tasks": (_evaluation_task(tmp_path, "task-a"),),
        "split": Split.TRAIN,
        "seed": 500,
        "repetitions": 1,
        "use_cache": True,
        "blind": False,
    }
    evaluator.evaluate(**kwargs)
    payload = json.loads(cache_path.read_text(encoding="utf-8"))
    result = next(iter(payload["rollout"].values()))["results"][0]
    result["reward"] = "2"
    result["success"] = "yes"
    result["raw_rewards"] = ["2"]
    cache_path.write_text(json.dumps(payload), encoding="utf-8")
    student.calls.clear()

    with pytest.raises(RuntimeError, match="incomplete cached evaluation bundle"):
        evaluator.cache_will_hit(
            **{key: value for key, value in kwargs.items() if key != "use_cache"}
        )
    with pytest.raises(RuntimeError, match="incomplete cached evaluation bundle"):
        evaluator.evaluate(**kwargs)
    assert student.calls == []


@pytest.mark.parametrize(
    ("field", "invalid_value", "task_id"),
    [
        ("task_id", 7, "7"),
        ("feedback", 7, "task-a"),
        ("evaluator_output", {}, "task-a"),
        ("final_answer", 7, "task-a"),
        ("visible_logs", [7], "task-a"),
    ],
)
def test_spreadsheet_evaluator_rejects_non_text_cached_fields(
    tmp_path: Path,
    field: str,
    invalid_value: object,
    task_id: str,
) -> None:
    cache_path = tmp_path / "rollouts.json"
    student = RecordingStudent()
    evaluator = evaluation.SpreadsheetSkillEvaluator(
        student,
        cache=JsonFileCache(cache_path),
        gate_metric="hard",
        gate_mixed_weight=0.5,
    )
    kwargs = {
        "skill": SkillArtifact("main", "main", "", "## Rule\nUpdate cells."),
        "tasks": (_evaluation_task(tmp_path, task_id),),
        "split": Split.TRAIN,
        "seed": 10,
        "repetitions": 1,
        "use_cache": True,
        "blind": False,
    }
    evaluator.evaluate(**kwargs)
    payload = json.loads(cache_path.read_text(encoding="utf-8"))
    result = next(iter(payload["rollout"].values()))["results"][0]
    result[field] = invalid_value
    cache_path.write_text(json.dumps(payload), encoding="utf-8")
    student.calls.clear()

    with pytest.raises(RuntimeError, match="incomplete cached evaluation bundle"):
        evaluator.evaluate(**kwargs)
    assert student.calls == []


@pytest.mark.parametrize(
    ("usage_field", "invalid_value"),
    [
        ("student_rollouts", "1"),
        ("input_tokens", True),
        ("cost_usd", "0.01"),
        ("elapsed_s", None),
    ],
)
def test_spreadsheet_evaluator_rejects_invalid_cached_usage_types(
    tmp_path: Path,
    usage_field: str,
    invalid_value: object,
) -> None:
    cache_path = tmp_path / "rollouts.json"
    student = RecordingStudent()
    evaluator = evaluation.SpreadsheetSkillEvaluator(
        student,
        cache=JsonFileCache(cache_path),
        gate_metric="hard",
        gate_mixed_weight=0.5,
    )
    kwargs = {
        "skill": SkillArtifact("main", "main", "", "## Rule\nUpdate cells."),
        "tasks": (_evaluation_task(tmp_path, "task-a"),),
        "split": Split.TRAIN,
        "seed": 10,
        "repetitions": 1,
        "use_cache": True,
        "blind": False,
    }
    evaluator.evaluate(**kwargs)
    payload = json.loads(cache_path.read_text(encoding="utf-8"))
    cached_batch = next(iter(payload["rollout"].values()))
    cached_batch["usage"][usage_field] = invalid_value
    cache_path.write_text(json.dumps(payload), encoding="utf-8")
    student.calls.clear()

    with pytest.raises(RuntimeError, match="incomplete cached evaluation bundle"):
        evaluator.evaluate(**kwargs)
    assert student.calls == []


def test_generic_task_mapping_keeps_strong_cache_identity(tmp_path: Path) -> None:
    cache = JsonFileCache(tmp_path / "rollouts.json")
    student = RecordingStudent()
    evaluator = evaluation.SpreadsheetSkillEvaluator(
        student,
        cache=cache,
        gate_metric="hard",
        gate_mixed_weight=0.5,
    )
    task = _evaluation_task(tmp_path, "task-a")
    changed_task = copy.deepcopy(task)
    changed_task["description"] = "Different content, same task id and files."
    kwargs = {
        "skill": SkillArtifact("main", "main", "", "## Rule\nUpdate cells."),
        "split": Split.TRAIN,
        "seed": 10,
        "repetitions": 1,
        "use_cache": True,
        "blind": False,
    }

    first = evaluator.evaluate(
        tasks=(MappingProxyType(task),),
        **kwargs,
    )
    changed = evaluator.evaluate(
        tasks=(MappingProxyType(changed_task),),
        **kwargs,
    )

    assert first.cache_hit is False
    assert changed.cache_hit is False
    assert len(student.calls) == 2


def test_blind_test_hides_answer_metadata_and_does_not_retry_from_score(
    tmp_path: Path,
    monkeypatch,
) -> None:
    client = RecordingClient()
    student = SpreadsheetStudent(STUDENT_CONFIG, client)
    monkeypatch.setattr(
        student.executor,
        "execute_and_score",
        lambda **kwargs: ExecutionResult(0.25, 1, 4, ""),
    )
    skill = SkillArtifact("main", "main", "", "## Rule\nUpdate cells.")

    trajectory = student.run_task(_final_task(tmp_path), skill, blind=True, seed=11)

    assert trajectory.hard_reward == 0.25
    assert len(client.calls) == 1
    prompt = repr(client.calls[0])
    assert "SECRET_FINAL_SHEET" not in prompt
    assert "B17:C19" not in prompt
    assert "score=" not in prompt
    assert client.calls[0]["seed"] == 11


def test_student_has_no_implicit_or_no_skill_entrypoints() -> None:
    assert not hasattr(SpreadsheetStudent, "run_task_no_skill")
    assert not hasattr(SpreadsheetStudent, "run_task_with_model")


def test_executor_copies_input_executes_code_and_scores_configured_range(
    tmp_path: Path,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, "input")
    _write_workbook(golden_file, "golden")
    code = (
        "from openpyxl import load_workbook\n"
        "wb = load_workbook(wb_path)\n"
        "wb['SECRET_FINAL_SHEET']['A1'] = 'golden'\n"
        "wb.save(wb_path)\n"
    )

    result = SpreadsheetExecutor().execute_and_score(
        code=code,
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position="A1",
        answer_sheet="SECRET_FINAL_SHEET",
    )

    assert result == ExecutionResult(score=1.0, matched=1, total=1)
    assert load_workbook(init_file)["SECRET_FINAL_SHEET"]["A1"].value == "input"


def test_executor_rejects_uncached_formula_values_instead_of_matching_none(
    tmp_path: Path,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, "=1+1")
    _write_workbook(golden_file, "=2+2")

    result = _execute_without_editing(init_file, golden_file)

    assert result == ExecutionResult(0.0, 0, 0, "missing_formula_cache")


def test_executor_matches_formula_cells_with_cached_empty_strings(
    tmp_path: Path,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    for path in (init_file, golden_file):
        _write_workbook(path, '=IF(1=1,"","")')
        _set_formula_cache_metadata(
            path,
            cell_type="str",
            include_value=True,
        )

    result = _execute_without_editing(init_file, golden_file)

    assert result == ExecutionResult(1.0, 1, 1)


@pytest.mark.parametrize(
    ("cell_type", "include_value"),
    [("str", False), ("n", True)],
)
def test_executor_rejects_missing_or_non_string_empty_formula_cache(
    tmp_path: Path,
    cell_type: str,
    include_value: bool,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    for path in (init_file, golden_file):
        _write_workbook(path, '=IF(1=1,"","")')
        _set_formula_cache_metadata(
            path,
            cell_type=cell_type,
            include_value=include_value,
        )

    result = _execute_without_editing(init_file, golden_file)

    assert result == ExecutionResult(0.0, 0, 0, "missing_formula_cache")


def test_executor_ignores_formula_metadata_outside_sheet_data(
    tmp_path: Path,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, '=IF(1=1,"","")')
    _set_formula_cache_metadata(
        init_file,
        cell_type="str",
        include_value=False,
        append_duplicate_outside_sheet_data=True,
    )
    _write_workbook(golden_file, '=IF(1=1,"","")')
    _set_formula_cache_metadata(
        golden_file,
        cell_type="str",
        include_value=True,
    )

    result = _execute_without_editing(init_file, golden_file)

    assert result == ExecutionResult(0.0, 0, 0, "missing_formula_cache")


def test_executor_ignores_formula_cache_from_foreign_xml_namespace(
    tmp_path: Path,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, '=IF(1=1,"","")')
    _set_formula_cache_metadata(
        init_file,
        cell_type="str",
        include_value=False,
        append_foreign_value=True,
    )
    _write_workbook(golden_file, '=IF(1=1,"","")')
    _set_formula_cache_metadata(
        golden_file,
        cell_type="str",
        include_value=True,
    )

    result = _execute_without_editing(init_file, golden_file)

    assert result == ExecutionResult(0.0, 0, 0, "missing_formula_cache")


def test_executor_ignores_sheet_mapping_outside_sheets(tmp_path: Path) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    workbook = Workbook()
    answer_sheet = workbook.active
    answer_sheet.title = "SECRET_FINAL_SHEET"
    answer_sheet["A1"] = '=IF(1=1,"","")'
    decoy_sheet = workbook.create_sheet("DECOY")
    decoy_sheet["A1"] = '=IF(1=1,"","")'
    workbook.save(init_file)
    _set_formula_cache_metadata(
        init_file,
        cell_type="str",
        include_value=False,
    )
    _set_formula_cache_metadata(
        init_file,
        cell_type="str",
        include_value=True,
        worksheet_name="xl/worksheets/sheet2.xml",
    )
    _append_fake_sheet_outside_sheets(
        init_file,
        fake_name="SECRET_FINAL_SHEET",
        target_sheet_name="DECOY",
    )
    _write_workbook(golden_file, '=IF(1=1,"","")')
    _set_formula_cache_metadata(
        golden_file,
        cell_type="str",
        include_value=True,
    )

    result = _execute_without_editing(init_file, golden_file)

    assert result == ExecutionResult(0.0, 0, 0, "missing_formula_cache")


def test_executor_rejects_duplicate_workbook_relationship_id(tmp_path: Path) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    for path in (init_file, golden_file):
        _write_workbook(path, '=IF(1=1,"","")')
        _set_formula_cache_metadata(
            path,
            cell_type="str",
            include_value=True,
        )
    _duplicate_workbook_relationship_id(
        init_file,
        sheet_name="SECRET_FINAL_SHEET",
    )

    result = _execute_without_editing(init_file, golden_file)

    assert result == ExecutionResult(0.0, 0, 0, "invalid_result_workbook")


def test_executor_distinguishes_boolean_from_numeric_values(tmp_path: Path) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, True)
    _write_workbook(golden_file, 1)

    result = _execute_without_editing(init_file, golden_file)

    assert result == ExecutionResult(0.0, 0, 1)


def test_executor_requires_matching_merge_topology_in_answer_range(
    tmp_path: Path,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, "same")
    _write_workbook(golden_file, "same")
    result_workbook = load_workbook(init_file)
    result_workbook["SECRET_FINAL_SHEET"].merge_cells("A1:B1")
    result_workbook.save(init_file)

    result = _execute_without_editing(
        init_file,
        golden_file,
        answer_position="A1:B1",
    )

    assert result == ExecutionResult(0.0, 0, 0, "merge_topology_mismatch")


def test_executor_compares_merge_anchor_when_range_contains_only_non_anchor(
    tmp_path: Path,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, "wrong")
    _write_workbook(golden_file, "right")
    for path in (init_file, golden_file):
        workbook = load_workbook(path)
        workbook["SECRET_FINAL_SHEET"].merge_cells("A1:B1")
        workbook.save(path)

    result = _execute_without_editing(
        init_file,
        golden_file,
        answer_position="B1",
    )

    assert result == ExecutionResult(0.0, 0, 1)


@pytest.mark.parametrize("unsupported_value", [float("nan"), "#DIV/0!"])
def test_executor_rejects_nonfinite_or_unsupported_cell_values(
    tmp_path: Path,
    unsupported_value: object,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, unsupported_value)
    _write_workbook(golden_file, unsupported_value)

    result = _execute_without_editing(init_file, golden_file)

    assert result == ExecutionResult(0.0, 0, 0, "unsupported_cell_value")


def test_executor_strictly_matches_supported_typed_values(tmp_path: Path) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    values = (17, 2.5, "exact", datetime(2026, 7, 15, 3, 30))
    for path in (init_file, golden_file):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "SECRET_FINAL_SHEET"
        for column, value in enumerate(values, start=1):
            worksheet.cell(row=1, column=column, value=value)
        workbook.save(path)

    result = _execute_without_editing(
        init_file,
        golden_file,
        answer_position="A1:D1",
    )

    assert result == ExecutionResult(1.0, 4, 4)


def test_executor_allows_matching_styled_blank_cells(tmp_path: Path) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    for path in (init_file, golden_file):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "SECRET_FINAL_SHEET"
        worksheet["A1"].fill = PatternFill(
            fill_type="solid",
            fgColor="FFFF00",
        )
        workbook.save(path)

    result = _execute_without_editing(init_file, golden_file)

    assert result == ExecutionResult(1.0, 1, 1)


def test_executor_rejects_empty_code_without_running_a_subprocess(
    tmp_path: Path,
    monkeypatch,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, "same")
    _write_workbook(golden_file, "same")
    monkeypatch.setattr(
        "rl_skill_edit.adapters.spreadsheet.subprocess.Popen",
        lambda *args, **kwargs: pytest.fail("empty code must not be executed"),
    )

    result = SpreadsheetExecutor().execute_and_score(
        code="   \n",
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position="A1",
        answer_sheet="SECRET_FINAL_SHEET",
    )

    assert result.score == 0.0
    assert result.total == 0
    assert result.error == "missing_executable_code"


@pytest.mark.parametrize(
    ("field", "replacement", "expected_error"),
    [
        ("init_file", "missing-input.xlsx", "missing_input_workbook"),
        ("golden_file", "missing-golden.xlsx", "missing_golden_workbook"),
        ("answer_position", "", "missing_answer_position"),
        ("answer_sheet", "", "missing_answer_sheet"),
    ],
)
def test_executor_preflight_failures_return_explicit_invalid_results(
    tmp_path: Path,
    monkeypatch,
    field: str,
    replacement: str,
    expected_error: str,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, "input")
    _write_workbook(golden_file, "golden")
    kwargs = {
        "code": "raise AssertionError('must not run')",
        "init_file": str(init_file),
        "golden_file": str(golden_file),
        "answer_position": "A1",
        "answer_sheet": "SECRET_FINAL_SHEET",
    }
    kwargs[field] = (
        str(tmp_path / replacement)
        if field in {"init_file", "golden_file"}
        else replacement
    )
    monkeypatch.setattr(
        "rl_skill_edit.adapters.spreadsheet.subprocess.Popen",
        lambda *args, **kwargs: pytest.fail("invalid inputs must not be executed"),
    )

    result = SpreadsheetExecutor().execute_and_score(**kwargs)

    assert result == ExecutionResult(0.0, 0, 0, expected_error)


def test_executor_timeout_is_invalid_and_uses_only_allowed_environment(
    tmp_path: Path,
    monkeypatch,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, "input")
    _write_workbook(golden_file, "golden")
    captured: dict = {"timeouts": []}

    class TimedOutProcess:
        pid = 12345
        returncode = None

        def communicate(self, *, timeout):
            captured["timeouts"].append(timeout)
            if len(captured["timeouts"]) == 1:
                raise subprocess.TimeoutExpired(captured["command"], timeout)
            self.returncode = -9
            return "", ""

    def fake_popen(command, **kwargs):
        captured["command"] = command
        captured["runner_source"] = Path(command[1]).read_text(encoding="utf-8")
        captured.update(kwargs)
        return TimedOutProcess()

    monkeypatch.setattr(
        "rl_skill_edit.adapters.spreadsheet.subprocess.Popen",
        fake_popen,
    )
    executor = SpreadsheetExecutor()
    killed: list[int] = []
    monkeypatch.setattr(
        executor,
        "_kill_process_tree",
        lambda process: killed.append(process.pid),
    )

    result = executor.execute_and_score(
        code="while True:\n    pass",
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position="A1",
        answer_sheet="SECRET_FINAL_SHEET",
    )

    assert result == ExecutionResult(0.0, 0, 0, "timeout_8s")
    assert captured["timeouts"] == [8, 1]
    assert killed == [12345]
    assert set(captured["env"]) == {
        "PATH",
        "PYTHONIOENCODING",
        "PYTHONDONTWRITEBYTECODE",
    }
    assert captured["start_new_session"] is (os.name != "nt")
    runner_source = captured["runner_source"]
    assert runner_source.index("resource.setrlimit") < runner_source.index(
        "probe_pid = os.fork()"
    )
    assert runner_source.index("probe_pid = os.fork()") < runner_source.index(
        "code_path ="
    )
    assert "os.waitpid(probe_pid, 0)" in runner_source


@pytest.mark.skipif(os.name == "nt", reason="RLIMIT_NPROC is POSIX-only")
@pytest.mark.parametrize("exit_mode", ["normal", "nonzero", "timeout"])
def test_executor_prevents_detached_descendants_for_every_exit_path(
    tmp_path: Path,
    exit_mode: str,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    child_pid_file = tmp_path / "detached-child.pid"
    blocked_file = tmp_path / "spawn-blocked.txt"
    _write_workbook(init_file, "same")
    _write_workbook(golden_file, "same")
    mode_action = {
        "normal": "pass",
        "nonzero": "raise RuntimeError('intentional failure')",
        "timeout": "time.sleep(30)",
    }[exit_mode]
    code = (
        "import subprocess, sys, time\n"
        "from pathlib import Path\n"
        "try:\n"
        "    child = subprocess.Popen(\n"
        "        [sys.executable, '-c', 'import time; time.sleep(30)'],\n"
        "        stdout=subprocess.DEVNULL,\n"
        "        stderr=subprocess.DEVNULL,\n"
        "        start_new_session=True,\n"
        "    )\n"
        f"    Path({str(child_pid_file)!r}).write_text(str(child.pid))\n"
        "except BlockingIOError as exc:\n"
        f"    Path({str(blocked_file)!r}).write_text(str(exc))\n"
        f"{mode_action}\n"
    )
    executor = SpreadsheetExecutor()
    if exit_mode == "timeout":
        executor._TIMEOUT_SECONDS = 0.2

    result = executor.execute_and_score(
        code=code,
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position="A1",
        answer_sheet="SECRET_FINAL_SHEET",
    )
    child_pid = (
        int(child_pid_file.read_text(encoding="utf-8"))
        if child_pid_file.exists()
        else None
    )
    try:
        assert blocked_file.is_file()
        assert child_pid is None
        if exit_mode == "normal":
            assert result == ExecutionResult(1.0, 1, 1)
        elif exit_mode == "nonzero":
            assert result.score == 0.0
            assert "intentional failure" in result.error
        else:
            assert result == ExecutionResult(0.0, 0, 0, "timeout_8s")
    finally:
        if child_pid is not None and _process_exists(child_pid):
            os.kill(child_pid, signal.SIGKILL)


@pytest.mark.skipif(os.name == "nt", reason="fork self-probe is POSIX-only")
def test_executor_runs_student_after_eagain_probe(tmp_path: Path, monkeypatch) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    student_marker = tmp_path / "student-ran.txt"
    _write_workbook(init_file, "same")
    _write_workbook(golden_file, "same")
    _install_runner_startup_hook(
        monkeypatch,
        _isolation_resource_source(fork_errno=errno.EAGAIN),
    )

    result = SpreadsheetExecutor().execute_and_score(
        code=(
            "from pathlib import Path\n"
            f"Path({str(student_marker)!r}).write_text('ran')\n"
        ),
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position="A1",
        answer_sheet="SECRET_FINAL_SHEET",
    )

    assert result == ExecutionResult(1.0, 1, 1)
    assert student_marker.read_text(encoding="utf-8") == "ran"


@pytest.mark.skipif(os.name == "nt", reason="fork self-probe is POSIX-only")
@pytest.mark.parametrize(
    "resource_source",
    [
        pytest.param(
            _isolation_resource_source(limits=(1, 0)),
            id="limits-not-zero",
        ),
        pytest.param(
            _isolation_resource_source(fork_errno=errno.ENOMEM),
            id="fork-enomem",
        ),
        pytest.param(
            _isolation_resource_source(uid=0),
            id="real-uid-root",
        ),
        pytest.param(
            _isolation_resource_source(euid=0),
            id="effective-uid-root",
        ),
        pytest.param(
            _isolation_resource_source(cap_eff=f"{1 << 21:x}"),
            id="cap-sys-admin",
        ),
        pytest.param(
            _isolation_resource_source(cap_eff=f"{1 << 24:x}"),
            id="cap-sys-resource",
        ),
        pytest.param(
            _isolation_resource_source(cap_eff="not-hex"),
            id="invalid-cap-eff",
        ),
    ],
)
def test_executor_fails_closed_when_process_isolation_is_unverified(
    tmp_path: Path,
    monkeypatch,
    resource_source: str,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    student_marker = tmp_path / "student-ran.txt"
    _write_workbook(init_file, "same")
    _write_workbook(golden_file, "same")
    _install_runner_startup_hook(monkeypatch, resource_source)

    result = SpreadsheetExecutor().execute_and_score(
        code=(
            "from pathlib import Path\n"
            f"Path({str(student_marker)!r}).write_text('ran')\n"
        ),
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position="A1",
        answer_sheet="SECRET_FINAL_SHEET",
    )

    assert "process_isolation_unavailable" in result.error
    assert not student_marker.exists()


@pytest.mark.skipif(os.name == "nt", reason="fork self-probe is POSIX-only")
def test_executor_fails_closed_and_reaps_probe_when_rlimit_is_ineffective(
    tmp_path: Path,
    monkeypatch,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    student_marker = tmp_path / "student-ran.txt"
    probe_pid_file = tmp_path / "probe.pid"
    waited_file = tmp_path / "probe.waited"
    _write_workbook(init_file, "same")
    _write_workbook(golden_file, "same")
    _install_runner_startup_hook(
        monkeypatch,
        _isolation_resource_source(fork_errno=None)
        + (
            "from pathlib import Path\n"
            "_real_fork = os.fork\n"
            "_real_waitpid = os.waitpid\n"
            "def _tracked_fork():\n"
            "    pid = _real_fork()\n"
            "    if pid > 0:\n"
            f"        Path({str(probe_pid_file)!r}).write_text(str(pid))\n"
            "    return pid\n"
            "def _tracked_waitpid(pid, options):\n"
            "    waited_pid, status = _real_waitpid(pid, options)\n"
            f"    with Path({str(waited_file)!r}).open('a') as handle:\n"
            "        handle.write(f'{pid}:{options}:{waited_pid}:{status}\\n')\n"
            "    return waited_pid, status\n"
            "os.fork = _tracked_fork\n"
            "os.waitpid = _tracked_waitpid\n"
        ),
    )

    result = SpreadsheetExecutor().execute_and_score(
        code=(
            "from pathlib import Path\n"
            f"Path({str(student_marker)!r}).write_text('ran')\n"
        ),
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position="A1",
        answer_sheet="SECRET_FINAL_SHEET",
    )

    assert "process_isolation_unavailable" in result.error
    assert not student_marker.exists()
    probe_pid = int(probe_pid_file.read_text(encoding="utf-8"))
    wait_calls = tuple(
        tuple(int(value) for value in line.split(":"))
        for line in waited_file.read_text(encoding="utf-8").splitlines()
    )
    assert len(wait_calls) == 1
    requested_pid, options, waited_pid, status = wait_calls[0]
    assert requested_pid == probe_pid
    assert options == 0
    assert waited_pid == probe_pid
    assert os.waitstatus_to_exitcode(status) == 0


@pytest.mark.skipif(os.name == "nt", reason="fork self-probe is POSIX-only")
@pytest.mark.parametrize(
    "function_name",
    ["getuid", "geteuid", "fork", "waitpid"],
)
def test_executor_fails_closed_when_posix_function_is_unavailable(
    tmp_path: Path,
    monkeypatch,
    function_name: str,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    student_marker = tmp_path / "student-ran.txt"
    _write_workbook(init_file, "same")
    _write_workbook(golden_file, "same")
    _install_runner_startup_hook(
        monkeypatch,
        _isolation_resource_source(
            missing_os_function=function_name,
        ),
    )

    result = SpreadsheetExecutor().execute_and_score(
        code=(
            "from pathlib import Path\n"
            f"Path({str(student_marker)!r}).write_text('ran')\n"
        ),
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position="A1",
        answer_sheet="SECRET_FINAL_SHEET",
    )

    assert "process_isolation_unavailable" in result.error
    assert not student_marker.exists()


def test_executor_does_not_insert_an_automatic_save(tmp_path: Path) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, "input")
    _write_workbook(golden_file, "golden")
    code_without_save = (
        "from openpyxl import load_workbook\n"
        "wb = load_workbook(wb_path)\n"
        "wb['SECRET_FINAL_SHEET']['A1'] = 'golden'\n"
    )

    result = SpreadsheetExecutor().execute_and_score(
        code=code_without_save,
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position="A1",
        answer_sheet="SECRET_FINAL_SHEET",
    )

    assert result == ExecutionResult(score=0.0, matched=0, total=1)


@pytest.mark.parametrize(
    "code",
    [
        "import module_that_does_not_exist_rl_skill_edit",
        "raise RuntimeError('explicit execution failure')",
    ],
)
def test_executor_import_or_runtime_failure_is_explicitly_invalid(
    tmp_path: Path,
    code: str,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, "input")
    _write_workbook(golden_file, "golden")

    result = SpreadsheetExecutor().execute_and_score(
        code=code,
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position="A1",
        answer_sheet="SECRET_FINAL_SHEET",
    )

    assert result.score == 0.0
    assert result.total == 0
    assert result.error


@pytest.mark.parametrize(
    ("answer_sheet", "answer_position", "expected_error"),
    [
        ("MISSING_SHEET", "A1", "missing_golden_sheet"),
        ("SECRET_FINAL_SHEET", "not-a-range", "invalid_answer_position"),
        ("SECRET_FINAL_SHEET", "B2:A1", "invalid_answer_position"),
        ("SECRET_FINAL_SHEET", "A1048577", "invalid_answer_position"),
        ("SECRET_FINAL_SHEET", "XFE1", "invalid_answer_position"),
    ],
)
def test_executor_rejects_misaligned_golden_metadata_before_execution(
    tmp_path: Path,
    monkeypatch,
    answer_sheet: str,
    answer_position: str,
    expected_error: str,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, "input")
    _write_workbook(golden_file, "golden")
    monkeypatch.setattr(
        "rl_skill_edit.adapters.spreadsheet.subprocess.Popen",
        lambda *args, **kwargs: pytest.fail("invalid metadata must not be executed"),
    )

    result = SpreadsheetExecutor().execute_and_score(
        code="raise AssertionError('must not run')",
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position=answer_position,
        answer_sheet=answer_sheet,
    )

    assert result == ExecutionResult(0.0, 0, 0, expected_error)


def test_executor_rejects_an_invalid_input_workbook_before_execution(
    tmp_path: Path,
    monkeypatch,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    init_file.write_bytes(b"not an xlsx workbook")
    _write_workbook(golden_file, "golden")
    monkeypatch.setattr(
        "rl_skill_edit.adapters.spreadsheet.subprocess.Popen",
        lambda *args, **kwargs: pytest.fail("invalid input must not be executed"),
    )

    result = SpreadsheetExecutor().execute_and_score(
        code="raise AssertionError('must not run')",
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position="A1",
        answer_sheet="SECRET_FINAL_SHEET",
    )

    assert result == ExecutionResult(0.0, 0, 0, "invalid_input_workbook")


@pytest.mark.parametrize(
    ("code", "expected_error"),
    [
        (
            "from pathlib import Path\nPath(wb_path).unlink()\n",
            "missing_result_workbook",
        ),
        (
            "from pathlib import Path\nPath(wb_path).write_bytes(b'not xlsx')\n",
            "invalid_result_workbook",
        ),
        (
            "from openpyxl import load_workbook\n"
            "wb = load_workbook(wb_path)\n"
            "wb.create_sheet('OTHER')\n"
            "del wb['SECRET_FINAL_SHEET']\n"
            "wb.save(wb_path)\n",
            "missing_result_sheet",
        ),
    ],
)
def test_executor_rejects_missing_or_misaligned_execution_artifacts(
    tmp_path: Path,
    code: str,
    expected_error: str,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, "input")
    _write_workbook(golden_file, "golden")

    result = SpreadsheetExecutor().execute_and_score(
        code=code,
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position="A1",
        answer_sheet="SECRET_FINAL_SHEET",
    )

    assert result == ExecutionResult(0.0, 0, 0, expected_error)


def test_executor_process_launch_failure_is_explicitly_invalid(
    tmp_path: Path,
    monkeypatch,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, "input")
    _write_workbook(golden_file, "golden")

    def fail_launch(*args, **kwargs):
        raise OSError("interpreter missing")

    monkeypatch.setattr(
        "rl_skill_edit.adapters.spreadsheet.subprocess.Popen",
        fail_launch,
    )

    result = SpreadsheetExecutor().execute_and_score(
        code="print('attempt')",
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position="A1",
        answer_sheet="SECRET_FINAL_SHEET",
    )

    assert result == ExecutionResult(
        0.0,
        0,
        0,
        "execution_error:interpreter missing",
    )


@pytest.mark.parametrize(
    "usage",
    [
        {"ok": True, "total_tokens": 1},
        {"ok": False, "total_tokens": 1, "cost_usd": 0.0},
    ],
)
def test_student_rejects_incomplete_or_failed_api_bundles_before_execution(
    tmp_path: Path,
    monkeypatch,
    usage: dict,
) -> None:
    client = RecordingClient(usage=usage)
    student = SpreadsheetStudent(STUDENT_CONFIG, client)
    monkeypatch.setattr(
        student.executor,
        "execute_and_score",
        lambda **kwargs: pytest.fail("an invalid API bundle must not execute"),
    )

    trajectory = student.run_task(
        _final_task(tmp_path),
        SkillArtifact("main", "main", "", "## Rule\nUpdate cells."),
        blind=True,
        seed=11,
    )

    assert trajectory.evaluation_valid is False
    assert trajectory.hard_reward == 0.0
    assert trajectory.invalid_reason.startswith("api:")


def test_student_prompt_has_exactly_one_forced_skill_path(
    tmp_path: Path,
    monkeypatch,
) -> None:
    client = RecordingClient()
    student = SpreadsheetStudent(STUDENT_CONFIG, client)
    monkeypatch.setattr(
        student.executor,
        "execute_and_score",
        lambda **kwargs: ExecutionResult(1.0, 1, 1),
    )
    skill = SkillArtifact(
        "main",
        "Only Active Skill",
        "",
        "## Unique Rule\nWrite the requested value.",
    )

    trajectory = student.run_task(_final_task(tmp_path), skill, blind=False, seed=19)

    assert trajectory.evaluation_valid is True
    assert len(client.calls) == 1
    call = client.calls[0]
    assert call["system"].count("[ACTIVE SKILL:") == 1
    assert "[ACTIVE SKILL: Only Active Skill]" in call["system"]
    assert skill.body in call["system"]
    assert "SECRET_FINAL_SHEET" in call["system"]
    assert "B17:C19" in call["system"]
    assert call["messages"] == [{"role": "user", "content": "Update the workbook."}]


@pytest.mark.parametrize(
    "response",
    [
        "",
        "I did not provide executable Python code.",
        "```python\nthis is not valid Python !!!\n```",
    ],
)
def test_student_empty_missing_or_invalid_code_is_invalid(
    tmp_path: Path,
    response: str,
) -> None:
    client = RecordingClient(response=response)
    student = SpreadsheetStudent(STUDENT_CONFIG, client)

    trajectory = student.run_task(
        _final_task(tmp_path),
        SkillArtifact("main", "main", "", "## Rule\nUpdate cells."),
        blind=True,
        seed=7,
    )

    assert len(client.calls) == 1
    assert trajectory.evaluation_valid is False
    assert trajectory.hard_reward == 0.0
    assert trajectory.invalid_reason


@pytest.mark.parametrize("language", ["bash", "json"])
def test_student_rejects_non_python_fenced_blocks_before_execution(
    tmp_path: Path,
    monkeypatch,
    language: str,
) -> None:
    client = RecordingClient(response=f"```{language}\n{{}}\n```")
    student = SpreadsheetStudent(STUDENT_CONFIG, client)
    monkeypatch.setattr(
        student.executor,
        "execute_and_score",
        lambda **kwargs: pytest.fail("non-Python code must not be executed"),
    )

    trajectory = student.run_task(
        _final_task(tmp_path),
        SkillArtifact("main", "main", "", "## Rule\nUpdate cells."),
        blind=True,
        seed=7,
    )

    assert trajectory.evaluation_valid is False
    assert trajectory.invalid_reason == "missing_executable_code"


def test_student_skips_non_python_fence_and_extracts_following_python(
    tmp_path: Path,
    monkeypatch,
) -> None:
    python_code = "workbook_value = 1"
    client = RecordingClient(
        response=(
            "```bash\n"
            "echo must-not-run\n"
            "```\n"
            "This explanation is not executable.\n"
            "```python\n"
            f"{python_code}\n"
            "```"
        )
    )
    student = SpreadsheetStudent(STUDENT_CONFIG, client)
    captured: dict[str, object] = {}

    def execute_and_score(**kwargs) -> ExecutionResult:
        captured.update(kwargs)
        return ExecutionResult(1.0, 1, 1)

    monkeypatch.setattr(student.executor, "execute_and_score", execute_and_score)

    trajectory = student.run_task(
        _final_task(tmp_path),
        SkillArtifact("main", "main", "", "## Rule\nUpdate cells."),
        blind=True,
        seed=7,
    )

    assert trajectory.evaluation_valid is True
    assert captured["code"] == python_code


def test_executor_injects_wb_path_without_preceding_future_import(
    tmp_path: Path,
) -> None:
    init_file = tmp_path / "input.xlsx"
    golden_file = tmp_path / "golden.xlsx"
    _write_workbook(init_file, "input")
    _write_workbook(golden_file, "golden")
    code = (
        "from __future__ import annotations\n"
        "from openpyxl import load_workbook\n"
        "workbook = load_workbook(wb_path)\n"
        "workbook['SECRET_FINAL_SHEET']['A1'] = 'golden'\n"
        "workbook.save(wb_path)\n"
    )

    result = SpreadsheetExecutor().execute_and_score(
        code=code,
        init_file=str(init_file),
        golden_file=str(golden_file),
        answer_position="A1",
        answer_sheet="SECRET_FINAL_SHEET",
    )

    assert result == ExecutionResult(1.0, 1, 1)


@pytest.mark.parametrize(
    "invalid_case",
    ["missing_description", "missing_answer_position", "missing_input_file"],
)
def test_student_rejects_an_incomplete_task_before_the_api_call(
    tmp_path: Path,
    invalid_case: str,
) -> None:
    task = _final_task(tmp_path)
    if invalid_case == "missing_description":
        del task["description"]
    elif invalid_case == "missing_answer_position":
        del task["spreadsheet"]["answer_position"]
    else:
        task["spreadsheet"]["init_file"] = str(tmp_path / "missing.xlsx")
    client = RecordingClient()
    student = SpreadsheetStudent(STUDENT_CONFIG, client)

    with pytest.raises(ValueError, match="task"):
        student.run_task(
            task,
            SkillArtifact("main", "main", "", "## Rule\nUpdate cells."),
            blind=True,
            seed=7,
        )

    assert client.calls == []
